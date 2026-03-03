"""
Complete extraction of Total Cost Summary from Excel - Check for missing items
"""
import openpyxl

SRC = r"C:\Users\dploy\OneDrive\Documents\Design to Basic Shirt Tool First Trail Version 1 - 20250801-update latest.xlsm"

wb = openpyxl.load_workbook(SRC, data_only=True, keep_vba=True)
ws_dl = wb['Data link']

print("=" * 80)
print("COMPLETE TOTAL COST SUMMARY - ALL ROWS")
print("=" * 80)

# Check rows 1-30 in columns AA, AB, AC
print("\nColumns AA, AB, AC (rows 1-30):")
print(f"{'Row':<5} {'AA (Label)':<50} {'AB (Value)':<20} {'AC (Value)':<20}")
print("-" * 100)

for row in range(1, 31):
    aa = ws_dl.cell(row, 27).value  # Column AA
    ab = ws_dl.cell(row, 28).value  # Column AB
    ac = ws_dl.cell(row, 29).value  # Column AC

    if aa or ab is not None or ac is not None:
        aa_str = str(aa)[:48] if aa else ''
        ab_str = str(ab)[:18] if ab is not None else ''
        ac_str = str(ac)[:18] if ac is not None else ''
        print(f"{row:<5} {aa_str:<50} {ab_str:<20} {ac_str:<20}")

print("\n" + "=" * 80)
print("Check for additional cost items in other columns")
print("=" * 80)

# Check if there are more columns with cost data
for col in range(24, 35):  # Columns X to AH
    col_letter = openpyxl.utils.get_column_letter(col)
    header = ws_dl.cell(1, col).value

    if header and ('cost' in str(header).lower() or 'total' in str(header).lower()):
        print(f"\nColumn {col_letter} ({header}):")
        for row in range(2, 15):
            val = ws_dl.cell(row, col).value
            if val is not None:
                label = ws_dl.cell(row, 27).value  # AA column label
                print(f"  Row {row}: {label} = {val}")

print("\n" + "=" * 80)
print("TOTAL ITEMS COUNT")
print("=" * 80)

# Count items
item_count = 0
for row in range(2, 20):
    aa = ws_dl.cell(row, 27).value
    if aa and aa != 'TOTAL COST SUMMARY':
        item_count += 1
        print(f"{item_count}. {aa}")

print(f"\nTotal items in Excel: {item_count}")
print("\nPlease compare with webapp and tell me what's missing!")
