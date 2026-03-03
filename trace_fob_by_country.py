"""
Trace FOB cost by country from Excel
"""
import openpyxl

SRC = r"C:\Users\dploy\OneDrive\Documents\Design to Basic Shirt Tool First Trail Version 1 - 20250801-update latest.xlsm"

wb = openpyxl.load_workbook(SRC, data_only=True, keep_vba=True)
ws_bs = wb['Basic Shirts Costing Tool']

print("=" * 80)
print("FOB COST BY COUNTRY (from Excel)")
print("=" * 80)

# Check rows 30-45 for country comparison
print(f"\n{'Row':<5} {'Country (D)':<20} {'Labour Cost (O)':<20} {'FOB Cost (AC)':<20}")
print("-" * 80)

for row in range(30, 46):
    country = ws_bs.cell(row, 4).value  # Column D
    labour_cost = ws_bs.cell(row, 15).value  # Column O
    fob_cost = ws_bs.cell(row, 29).value  # Column AC

    if country:
        country_str = str(country)[:18] if country else ''
        labour_str = f"${labour_cost:.6f}" if isinstance(labour_cost, (int, float)) else str(labour_cost)[:18]
        fob_str = f"${fob_cost:.6f}" if isinstance(fob_cost, (int, float)) else str(fob_cost)[:18]
        print(f"{row:<5} {country_str:<20} {labour_str:<20} {fob_str:<20}")

# Now check the formula to understand the calculation
print("\n" + "=" * 80)
print("FORMULA STRUCTURE")
print("=" * 80)

wb2 = openpyxl.load_workbook(SRC, data_only=False, keep_vba=True)
ws_bs2 = wb2['Basic Shirts Costing Tool']

print("\nFor BANGLADESH (Row 31):")
print(f"  Country: {ws_bs2.cell(31, 4).value}")
print(f"  Labour Cost Formula (O31): {ws_bs2.cell(31, 15).value}")
print(f"  FOB Cost Formula (AC31): {ws_bs2.cell(31, 29).value}")

# Check Data link sheet for the source
print("\n" + "=" * 80)
print("DATA LINK SOURCE")
print("=" * 80)

ws_dl = wb['Data link']
print("\nManufacturing rows (39-45):")
for row in range(39, 50):
    country = ws_dl.cell(row, 17).value  # Column Q
    minutes = ws_dl.cell(row, 19).value  # Column S
    efficiency = ws_dl.cell(row, 21).value  # Column U
    cost_rate = ws_dl.cell(row, 26).value  # Column Z
    total = ws_dl.cell(row, 28).value  # Column AB

    if country:
        print(f"\nRow {row}: {country}")
        print(f"  Minutes: {minutes}")
        print(f"  Efficiency: {efficiency}")
        print(f"  Cost Rate: {cost_rate}")
        print(f"  Total Cost: {total}")
