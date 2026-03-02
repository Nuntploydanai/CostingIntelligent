"""
Check efficiency lookup columns AA and AB
"""
import openpyxl

SRC = r"C:\Users\dploy\OneDrive\Documents\Design to Basic Shirt Tool First Trail Version 1 - 20250801-update latest.xlsm"

wb = openpyxl.load_workbook(SRC, data_only=True, keep_vba=True)
ws_sam = wb['SAM&Product EFF%']
ws_dl = wb['Data link']

print("=" * 80)
print("EFFICIENCY LOOKUP DATA (Columns AA and AB)")
print("=" * 80)

print("\nColumn AA and AB data:")
for row in range(1, 20):
    aa_val = ws_sam.cell(row, 27).value  # Column AA
    ab_val = ws_sam.cell(row, 28).value  # Column AB

    if aa_val is not None or ab_val is not None:
        print(f"Row {row}: AA={aa_val}, AB={ab_val}")

print("\n" + "=" * 80)
print("What's in E31 (quantity lookup key)?")
print("=" * 80)

print(f"E31: {ws_dl['E31'].value}")

print("\n" + "=" * 80)
print("What's in D22 (Ideal Quantity)?")
print("=" * 80)

ws_bs = wb['Basic Shirts Costing Tool']
print(f"D22: {ws_bs['D22'].value}")
