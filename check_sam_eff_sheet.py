"""
Check SAM&Product EFF% sheet with correct inputs
"""
import openpyxl

SRC = r"C:\Users\dploy\Downloads\Design to Basic Shirt Tool First Trail Version 1 - 20250801-update latest.xlsm"

wb = openpyxl.load_workbook(SRC, data_only=True, keep_vba=True)

ws_sam = wb['SAM&Product EFF%']
ws_dl = wb['Data link']
ws_bs = wb['Basic Shirts Costing Tool']

print("=" * 80)
print("INPUTS FROM YOUR SCREENSHOT")
print("=" * 80)

silhouette = ws_bs['D14'].value
seam = ws_bs['D15'].value
ideal_qty = ws_bs['D22'].value

print(f"D14 (Silhouette): {silhouette}")
print(f"D15 (Seam): {seam}")
print(f"D22 (Ideal Quantity): {ideal_qty}")

print("\n" + "=" * 80)
print("SAM&Product EFF% SHEET - Look for silhouette+seam combination")
print("=" * 80)

# Find the row that matches silhouette + seam
for row in range(1, 100):
    aa_val = ws_sam.cell(row, 27).value  # AA column
    ab_val = ws_sam.cell(row, 28).value  # AB column
    if aa_val and str(aa_val).strip() == str(ideal_qty).strip():
        print(f"Row {row}: AA={aa_val}, AB={ab_val}")

print("\n" + "=" * 80)
print("DATA LINK - E31 (Ideal Quantity)")
print("=" * 80)
print(f"E31: {ws_dl['E31'].value}")

print("\n" + "=" * 80)
print("DATA LINK - Q30 Formula and Value")
print("=" * 80)
wb2 = openpyxl.load_workbook(SRC, data_only=False, keep_vba=True)
ws_dl2 = wb2['Data link']
print(f"Q30 Formula: {ws_dl2['Q30'].value}")
print(f"Q30 Value: {ws_dl['Q30'].value}")

print("\n" + "=" * 80)
print("DATA LINK - J30 (Base Minutes)")
print("=" * 80)
print(f"J30 Value: {ws_dl['J30'].value}")

print("\n" + "=" * 80)
print("EXPECTED W18 VALUE FROM YOUR SCREENSHOT: 2.3")
print("WHAT I'M GETTING: 4.3")
print("WHERE DOES 2.3 COME FROM???")
print("=" * 80)
