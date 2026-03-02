"""
Trace AC column overhead costs
"""
import openpyxl

SRC = r"C:\Users\dploy\Downloads\Design to Basic Shirt Tool First Trail Version 1 - 20250801-update latest.xlsm"

wb = openpyxl.load_workbook(SRC, data_only=True, keep_vba=True)
ws_bs = wb['Basic Shirts Costing Tool']

print("=" * 80)
print("OVERHEAD COSTS (AC Column) - All rows")
print("=" * 80)

for row in range(1, 20):
    label = ws_bs.cell(row, 27).value  # AA column
    value = ws_bs.cell(row, 29).value  # AC column
    if label is not None or value is not None:
        print(f"AC{row} ({label}): {value}")

print("\n" + "=" * 80)
print("Formula for AC2-AC11 (from earlier trace):")
print("Z39 = IF($E$32 = X39, $AC$13,")
print("   ($AC$2+$AC$3+$AC$4+$AC$5+$AC$6+$AC$7+$AC$9+$AC$10+$AC$11+V39+W39))")
print("=" * 80)

# Calculate sum
overhead_sum = 0
for row in [2, 3, 4, 5, 6, 7, 9, 10, 11]:
    value = ws_bs.cell(row, 29).value
    if value is not None and isinstance(value, (int, float)):
        overhead_sum += value
        print(f"AC{row}: {value}")

print(f"\nSum of overheads: {overhead_sum}")
print(f"Plus V39+W39: need to trace these")
