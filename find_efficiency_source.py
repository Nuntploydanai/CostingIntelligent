"""
Find efficiency data source
"""
import openpyxl

SRC = r"C:\Users\dploy\OneDrive\Documents\Design to Basic Shirt Tool First Trail Version 1 - 20250801-update latest.xlsm"

wb_f = openpyxl.load_workbook(SRC, data_only=False, keep_vba=True)
wb_v = openpyxl.load_workbook(SRC, data_only=True, keep_vba=True)

ws_dl_f = wb_f['Data link']
ws_dl_v = wb_v['Data link']

print("=" * 80)
print("EFFICIENCY FORMULA")
print("=" * 80)

print(f"\nQ30 formula: {ws_dl_f['Q30'].value}")
print(f"Q30 value: {ws_dl_v['Q30'].value}")

print(f"\nN30 value: {ws_dl_v['N30'].value}")

# Check what N30 references
print(f"\nN30 formula: {ws_dl_f['N30'].value}")

# Check if there's a reference to SAM&Product EFF% sheet
print("\n" + "=" * 80)
print("Looking for efficiency data in SAM&Product EFF% sheet")
print("=" * 80)

ws_sam = wb_v['SAM&Product EFF%']

# Check column headers
print("\nColumn headers (Row 1):")
for col in range(1, 30):
    val = ws_sam.cell(1, col).value
    if val:
        col_letter = openpyxl.utils.get_column_letter(col)
        print(f"{col_letter}1: {val}")

# Check if there's an EFF% column
print("\nChecking column G (EFF%?):")
for row in range(1, 10):
    val = ws_sam.cell(row, 7).value
    if val:
        print(f"G{row}: {val}")
