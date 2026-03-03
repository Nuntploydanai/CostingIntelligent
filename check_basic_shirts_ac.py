"""
Check Basic Shirts Costing Tool column AC for actual values
"""
import openpyxl

SRC = r"C:\Users\dploy\OneDrive\Documents\Design to Basic Shirt Tool First Trail Version 1 - 20250801-update latest.xlsm"

wb = openpyxl.load_workbook(SRC, data_only=True, keep_vba=True)
ws_bs = wb['Basic Shirts Costing Tool']

print("=" * 80)
print("BASIC SHIRTS COSTING TOOL - COLUMN AC (ACTUAL COSTS)")
print("=" * 80)

print(f"\n{'Row':<5} {'AA (Label)':<50} {'AC (Value)':<20}")
print("-" * 80)

for row in range(6, 20):
    aa_label = ws_bs.cell(row, 27).value  # Column AA
    ac_value = ws_bs.cell(row, 29).value  # Column AC

    if aa_label or ac_value is not None:
        aa_str = str(aa_label)[:48] if aa_label else ''
        ac_str = str(ac_value)[:18] if ac_value is not None else ''
        print(f"{row:<5} {aa_str:<50} {ac_str:<20}")

print("\n" + "=" * 80)
print("CHECKING SUPPLIER MARGIN INPUT")
print("=" * 80)

# Look for where user inputs Supplier Margin
for row in range(1, 100):
    for col in range(1, 20):
        val = ws_bs.cell(row, col).value
        if val and isinstance(val, str) and 'margin' in val.lower():
            col_letter = openpyxl.utils.get_column_letter(col)
            print(f"\nFound at {col_letter}{row}: {val}")

            # Check if next column has a dropdown or input
            next_col_val = ws_bs.cell(row, col + 1).value
            print(f"  Next cell value: {next_col_val}")

# Check if there's a dropdown for margin percentage
print("\n" + "=" * 80)
print("LOOKING FOR MARGIN PERCENTAGE INPUT")
print("=" * 80)

# Check row 16 (where AC16 is referenced for margin)
print(f"\nRow 16:")
for col in range(1, 35):
    val = ws_bs.cell(16, col).value
    if val is not None:
        col_letter = openpyxl.utils.get_column_letter(col)
        print(f"  {col_letter}16: {val}")
