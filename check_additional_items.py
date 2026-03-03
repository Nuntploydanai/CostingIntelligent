"""
Check Excel for items 13-20 in Total Cost Summary
"""
import openpyxl

SRC = r"C:\Users\dploy\OneDrive\Documents\Design to Basic Shirt Tool First Trail Version 1 - 20250801-update latest.xlsm"

wb = openpyxl.load_workbook(SRC, data_only=True, keep_vba=True)
ws_dl = wb['Data link']

print("=" * 80)
print("TOTAL COST SUMMARY - ITEMS 13-20")
print("=" * 80)

print(f"\n{'Row':<5} {'Label (AA)':<50} {'Value (AC)':<20}")
print("-" * 80)

for row in range(2, 25):
    label = ws_dl.cell(row, 27).value  # Column AA
    value = ws_dl.cell(row, 29).value  # Column AC

    if label or value is not None:
        label_str = str(label)[:48] if label else ''
        value_str = str(value)[:18] if value is not None else ''
        print(f"{row:<5} {label_str:<50} {value_str:<20}")

# Check formulas for the new items
print("\n" + "=" * 80)
print("FORMULAS FOR ADDITIONAL ITEMS")
print("=" * 80)

wb2 = openpyxl.load_workbook(SRC, data_only=False, keep_vba=True)
ws_dl2 = wb2['Data link']

for row in range(13, 25):
    label = ws_dl2.cell(row, 27).value  # Column AA
    formula = ws_dl2.cell(row, 29).value  # Column AC

    if label:
        print(f"\nRow {row}: {label}")
        print(f"  Formula: {formula}")
