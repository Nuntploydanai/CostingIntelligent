"""
Find where AC column values (actual costs) come from
"""
import openpyxl

SRC = r"C:\Users\dploy\OneDrive\Documents\Design to Basic Shirt Tool First Trail Version 1 - 20250801-update latest.xlsm"

wb = openpyxl.load_workbook(SRC, data_only=False, keep_vba=True)
ws_dl = wb['Data link']

print("=" * 80)
print("TRACING ACTUAL COST VALUES (COLUMN AC)")
print("=" * 80)

print("\nChecking formulas for AC column (actual dollar values):")

for row in range(2, 16):
    label = ws_dl.cell(row, 27).value  # AA column
    if not label:
        continue

    # Get formula from AC
    ac_formula = ws_dl.cell(row, 29).value  # Column AC

    print(f"\nRow {row}: {label}")
    print(f"  AC formula: {ac_formula}")

# Check where Sewing Thread cost comes from
print("\n" + "=" * 80)
print("CHECKING SEWING THREAD COST SOURCE")
print("=" * 80)

# Check if there's a separate sewing thread calculation
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for row in range(1, 100):
        for col in range(1, 30):
            val = ws.cell(row, col).value
            if val and isinstance(val, str) and 'thread' in val.lower():
                col_letter = openpyxl.utils.get_column_letter(col)
                print(f"\n{sheet_name}!{col_letter}{row}: {val}")

# Check where Product Testing Cost comes from
print("\n" + "=" * 80)
print("CHECKING PRODUCT TESTING COST SOURCE")
print("=" * 80)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for row in range(1, 100):
        for col in range(1, 30):
            val = ws.cell(row, col).value
            if val and isinstance(val, str) and 'test' in val.lower():
                col_letter = openpyxl.utils.get_column_letter(col)
                print(f"\n{sheet_name}!{col_letter}{row}: {val}")
