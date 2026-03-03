"""
Trace where Total Cost Summary values come from in Excel
"""
import openpyxl

SRC = r"C:\Users\dploy\OneDrive\Documents\Design to Basic Shirt Tool First Trail Version 1 - 20250801-update latest.xlsm"

wb = openpyxl.load_workbook(SRC, data_only=True, keep_vba=True)
ws_dl = wb['Data link']

print("=" * 80)
print("TRACING TOTAL COST SUMMARY VALUES")
print("=" * 80)

# Check column AB (values) and their formulas
print("\nChecking formulas for each item:")

for row in range(2, 16):
    label = ws_dl.cell(row, 27).value  # AA column
    if not label:
        continue

    # Get value from AB
    ab_value = ws_dl.cell(row, 28).value

    # Get value from AC (calculated value)
    ac_value = ws_dl.cell(row, 29).value

    print(f"\nRow {row}: {label}")
    print(f"  AB (input/value): {ab_value}")
    print(f"  AC (calculated): {ac_value}")

# Check if Supplier Margin is a user input or calculated
print("\n" + "=" * 80)
print("CHECKING SUPPLIER MARGIN")
print("=" * 80)

# Look for where supplier margin is defined
wb_formulas = openpyxl.load_workbook(SRC, data_only=False, keep_vba=True)
ws_dl_f = wb_formulas['Data link']

print("\nRow 12 (% FOR SUPPLIER MARGIN):")
print(f"  AB12 formula: {ws_dl_f.cell(12, 28).value}")
print(f"  AB12 value: {ws_dl.cell(12, 28).value}")

# Check Basic Shirts for user inputs
ws_bs = wb['Basic Shirts Costing Tool']

print("\nLooking for Supplier Margin input in Basic Shirts...")
for row in range(1, 100):
    for col in range(1, 20):
        val = ws_bs.cell(row, col).value
        if val and 'margin' in str(val).lower():
            col_letter = openpyxl.utils.get_column_letter(col)
            print(f"  {col_letter}{row}: {val}")

            # Check if next cell has a value
            next_val = ws_bs.cell(row, col + 1).value
            if next_val is not None:
                print(f"    Value: {next_val}")

print("\n" + "=" * 80)
print("CHECKING PRODUCT TESTING COST")
print("=" * 80)

print("\nRow 9 (Total Product Testing Cost):")
print(f"  AB9 formula: {ws_dl_f.cell(9, 28).value}")
print(f"  AB9 value: {ws_dl.cell(9, 28).value}")
print(f"  AC9 value: {ws_dl.cell(9, 29).value}")

print("\n" + "=" * 80)
print("CHECKING SEWING THREAD COST")
print("=" * 80)

print("\nRow 7 (Total Sewing Thread cost):")
print(f"  AB7 formula: {ws_dl_f.cell(7, 28).value}")
print(f"  AB7 value: {ws_dl.cell(7, 28).value}")
print(f"  AC7 value: {ws_dl.cell(7, 29).value}")

print("\n" + "=" * 80)
print("CHECKING OTHER COST")
print("=" * 80)

print("\nRow 11 (Total other cost):")
print(f"  AB11 formula: {ws_dl_f.cell(11, 28).value}")
print(f"  AB11 value: {ws_dl.cell(11, 28).value}")
print(f"  AC11 value: {ws_dl.cell(11, 29).value}")

print("\n" + "=" * 80)
print("CHECKING PRINT/EMBROIDERY COST")
print("=" * 80)

print("\nRow 10 (Total print/embroidery cost):")
print(f"  AB10 formula: {ws_dl_f.cell(10, 28).value}")
print(f"  AB10 value: {ws_dl.cell(10, 28).value}")
print(f"  AC10 value: {ws_dl.cell(10, 29).value}")
