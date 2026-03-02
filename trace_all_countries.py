"""
Trace manufacturing cost table for ALL countries (rows 30-35)
"""
import openpyxl

SRC = r"C:\Users\dploy\Downloads\Design to Basic Shirt Tool First Trail Version 1 - 20250801-update latest.xlsm"

wb_f = openpyxl.load_workbook(SRC, data_only=False, keep_vba=True)
wb_v = openpyxl.load_workbook(SRC, data_only=True, keep_vba=True)

ws_bs_f = wb_f['Basic Shirts Costing Tool']
ws_bs_v = wb_v['Basic Shirts Costing Tool']
ws_dl_f = wb_f['Data link']
ws_dl_v = wb_v['Data link']

print("=" * 80)
print("MANUFACTURING COST TABLE - ALL COUNTRIES (Rows 30-35)")
print("=" * 80)

print("\n=== BASIC SHIRTS - Manufacturing Cost Rows ===")
for row in range(30, 36):
    country = ws_bs_v.cell(row, 4).value  # D column
    print(f"\nRow {row}: {country}")
    print(f"  D{row} (Country): {country}")

    # Get formulas and values for I, M, K, O columns
    for col, label in [(9, 'I-Minutes'), (13, 'K-CostRate'), (15, 'M-Efficiency'), (29, 'O-Total')]:
        col_letter = openpyxl.utils.get_column_letter(col)
        formula = ws_bs_f.cell(row, col).value
        value = ws_bs_v.cell(row, col).value
        print(f"  {col_letter}{row} ({label}): {value}")
        print(f"    Formula: {formula}")

print("\n" + "=" * 80)
print("DATA LINK - Source data for each country (rows 39-44)")
print("=" * 80)

for row in range(39, 45):
    country = ws_dl_v.cell(row, 12).value  # L column
    print(f"\nRow {row}: {country}")

    # Key columns
    for col, label in [(17, 'Q-Minutes'), (19, 'S-CostRate'), (21, 'U-Efficiency'), (26, 'Z-Total')]:
        col_letter = openpyxl.utils.get_column_letter(col)
        value = ws_dl_v.cell(row, col).value
        print(f"  {col_letter}{row} ({label}): {value}")

print("\n" + "=" * 80)
print("FOCUS: BANGLADESH (Row 31 in Basic Shirts, Row 40 in Data link)")
print("=" * 80)

# Bangladesh details
print("\nBANGLADESH in Basic Shirts (Row 31):")
print(f"  I31 (Minutes): {ws_bs_v.cell(31, 9).value}")
print(f"  K31 (Cost Rate): {ws_bs_v.cell(31, 13).value}")
print(f"  M31 (Efficiency): {ws_bs_v.cell(31, 15).value}")
print(f"  O31 (Total): {ws_bs_v.cell(31, 29).value}")

print("\nBANGLADESH in Data Link (Row 40):")
print(f"  Q40 (Minutes): {ws_dl_v.cell(40, 17).value}")
print(f"  S40 (Cost Rate): {ws_dl_v.cell(40, 19).value}")
print(f"  U40 (Efficiency): {ws_dl_v.cell(40, 21).value}")
print(f"  Z40 (Total): {ws_dl_v.cell(40, 26).value}")

print("\n" + "=" * 80)
print("Formulas for BANGLADESH:")
print("=" * 80)
print(f"I31: {ws_bs_f.cell(31, 9).value}")
print(f"K31: {ws_bs_f.cell(31, 13).value}")
print(f"M31: {ws_bs_f.cell(31, 15).value}")
print(f"O31: {ws_bs_f.cell(31, 29).value}")
