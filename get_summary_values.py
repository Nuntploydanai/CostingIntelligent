"""
Get Total Cost Summary values from Data link sheet
"""
import openpyxl

SRC = r"C:\Users\dploy\OneDrive\Documents\Design to Basic Shirt Tool First Trail Version 1 - 20250801-update latest.xlsm"

wb = openpyxl.load_workbook(SRC, data_only=True, keep_vba=True)
ws_dl = wb['Data link']

print("=" * 80)
print("TOTAL COST SUMMARY - FROM DATA LINK SHEET")
print("=" * 80)

# Column AA has labels, Column AB probably has values
summary_items = []

for row in range(1, 15):
    aa_label = ws_dl.cell(row, 27).value  # Column AA
    ab_value = ws_dl.cell(row, 28).value  # Column AB (values)

    if aa_label:
        print(f"\nRow {row}:")
        print(f"  Label (AA): {aa_label}")
        print(f"  Value (AB): {ab_value}")

        summary_items.append({
            'label': aa_label,
            'value': ab_value
        })

print("\n" + "=" * 80)
print("TOTAL COST SUMMARY STRUCTURE")
print("=" * 80)

print("\nWhat the webapp Total Cost Summary should show:")
for i, item in enumerate(summary_items, 1):
    print(f"{i}. {item['label']}: {item['value']}")

# Check if there's a total row
print("\n" + "=" * 80)
print("Looking for GRAND TOTAL...")
print("=" * 80)

for row in range(1, 20):
    aa_val = ws_dl.cell(row, 27).value
    ab_val = ws_dl.cell(row, 28).value
    ac_val = ws_dl.cell(row, 29).value

    if aa_val and 'total' in str(aa_val).lower():
        print(f"\nRow {row}:")
        print(f"  AA: {aa_val}")
        print(f"  AB: {ab_val}")
        print(f"  AC: {ac_val}")
