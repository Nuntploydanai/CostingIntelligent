"""
Find which combination gives exactly 2.3 minutes
"""
import openpyxl

SRC = r"C:\Users\dploy\Downloads\Design to Basic Shirt Tool First Trail Version 1 - 20250801-update latest.xlsm"

wb = openpyxl.load_workbook(SRC, data_only=True, keep_vba=True)

ws_sam = wb['SAM&Product EFF%']

print("=" * 80)
print("FIND ALL COMBINATIONS WITH 2.3 MINUTES")
print("=" * 80)

for row in range(1, 100):
    a_val = ws_sam.cell(row, 1).value  # Gender
    b_val = ws_sam.cell(row, 2).value  # Product shape
    c_val = ws_sam.cell(row, 3).value  # Seam
    d_val = ws_sam.cell(row, 4).value  # Size
    e_val = ws_sam.cell(row, 5).value  # SAM (minutes)

    if e_val is not None:
        try:
            minutes = float(e_val)
            if abs(minutes - 2.3) < 0.01:  # Check if close to 2.3
                print(f"\nRow {row}: Minutes = {minutes}")
                print(f"  Gender: {a_val}")
                print(f"  Product: {b_val}")
                print(f"  Seam: {c_val}")
                print(f"  Size: {d_val}")
        except:
            pass

print("\n" + "=" * 80)
print("Which combination matches your inputs?")
print("=" * 80)
