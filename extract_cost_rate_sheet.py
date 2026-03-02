"""
Extract all countries from " Cost Rate" sheet
"""
import openpyxl
import csv
from pathlib import Path

SRC = r"C:\Users\dploy\OneDrive\Documents\Design to Basic Shirt Tool First Trail Version 1 - 20250801-update latest.xlsm"
OUT_DIR = Path(r"C:\Users\dploy\.openclaw\workspace\basicshirts_web\master_clean")

wb = openpyxl.load_workbook(SRC, data_only=True, keep_vba=True)
ws_cr = wb[' Cost Rate']

print("=" * 80)
print("COST RATE SHEET DATA")
print("=" * 80)

cost_rate_data = []

for row in range(2, 50):
    country = ws_cr.cell(row, 1).value  # Column A
    cost_rate = ws_cr.cell(row, 15).value  # Column O

    if country and cost_rate is not None:
        cost_rate_data.append({
            'country': str(country).strip(),
            'cost_rate': float(cost_rate)
        })
        print(f"{country}: {cost_rate}")

# Write to CSV
csv_path = OUT_DIR / "cost_rate.csv"
with csv_path.open("w", newline="", encoding="utf-8") as f:
    w = csv.DictWriter(f, fieldnames=['country', 'cost_rate'])
    w.writeheader()
    for row in cost_rate_data:
        w.writerow(row)

print(f"\n\nWrote {len(cost_rate_data)} countries to {csv_path}")
