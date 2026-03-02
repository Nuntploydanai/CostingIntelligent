import csv
import re
from pathlib import Path

import openpyxl

SRC = r"C:\Users\dploy\Downloads\Design to Basic Shirt Tool First Trail Version 1 - 20250801-update latest.xlsm"
OUT_DIR = Path(r"C:\Users\dploy\.openclaw\workspace\basicshirts_web\master_clean")
OUT_DIR.mkdir(parents=True, exist_ok=True)

wb = openpyxl.load_workbook(SRC, data_only=False, keep_vba=True)
ws = wb["Conversion"]

# Cells known to contain fabric GSM default logic based on Data link H7/H8/H9
cells = ["B3", "B28", "B53"]

# Extract all string literals compared in IF(... H7="TYPE" ...)
# Simple regex for "..." strings
str_re = re.compile(r'"([^"]*)"')

fabric_types = set()
for addr in cells:
    f = ws[addr].value
    if not isinstance(f, str):
        continue
    for s in str_re.findall(f):
        s = (s or "").strip()
        if s in ("", "-", "0"):
            continue
        # keep simple word-ish tokens only
        if not re.fullmatch(r"[A-Za-z0-9]+(?:[xX][0-9]+)?", s):
            continue
        fabric_types.add(s)

# This grabs lots of literals; filter to those that look like fabric types
# We keep those that appear in Fabric Price sheet somewhere.
ws_fp = wb["Fabric Price"]
fp_text = set()
for row in ws_fp.iter_rows():
    for c in row:
        v = c.value
        if isinstance(v, str) and v.strip():
            fp_text.add(v.strip())

final = sorted([t for t in fabric_types if t in fp_text])

out_path = OUT_DIR / "dropdown_fabric_type.csv"
with out_path.open("w", newline="", encoding="utf-8") as f:
    w = csv.writer(f)
    w.writerow(["value"])
    for t in final:
        w.writerow([t])

print(f"wrote {out_path} count={len(final)}")
