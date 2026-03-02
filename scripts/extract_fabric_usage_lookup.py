import csv
from pathlib import Path
import openpyxl

SRC = r"C:\Users\dploy\Downloads\Design to Basic Shirt Tool First Trail Version 1 - 20250801-update latest.xlsm"
OUT = Path(r"C:\Users\dploy\.openclaw\workspace\basicshirts_web\master_clean\fabric_usage_lookup.csv")

wb = openpyxl.load_workbook(SRC, data_only=True, keep_vba=True)
ws = wb["Fabric Usage"]

# Data link R7: XLOOKUP(K7&L7, Fabric Usage!C:C&Fabric Usage!D:D, Fabric Usage!E:E, "X")
# So we extract from columns C,D,E.
rows = []
for r in range(1, ws.max_row + 1):
    c = ws.cell(r, 3).value
    d = ws.cell(r, 4).value
    e = ws.cell(r, 5).value

    if c is None and d is None and e is None:
        continue

    c_s = (str(c).strip() if c is not None else "")
    d_s = (str(d).strip() if d is not None else "")

    # usage should be numeric; allow strings that parse
    if e is None or e == "":
        continue
    try:
        usage = float(e)
    except Exception:
        # skip non-numeric rows
        continue

    # heuristics: skip header-like rows
    if c_s.lower() in ("type", "silhouette", "", "-") and d_s.lower() in ("", "-"):
        continue

    if not c_s or not d_s:
        continue

    rows.append((c_s, d_s, usage))

# de-dup: keep first
seen = set()
uniq = []
for c_s, d_s, usage in rows:
    k = (c_s, d_s)
    if k in seen:
        continue
    seen.add(k)
    uniq.append((c_s, d_s, usage))

OUT.parent.mkdir(parents=True, exist_ok=True)
with OUT.open("w", newline="", encoding="utf-8") as f:
    w = csv.writer(f)
    w.writerow(["k7_key", "using_part", "usage"])
    for c_s, d_s, usage in uniq:
        w.writerow([c_s, d_s, usage])

print(f"wrote {OUT} rows={len(uniq)}")
