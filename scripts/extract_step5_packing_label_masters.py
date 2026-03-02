import csv
from pathlib import Path
import openpyxl

SRC = r"C:\Users\dploy\Downloads\Design to Basic Shirt Tool First Trail Version 1 - 20250801-update latest.xlsm"
OUT_DIR = Path(r"C:\Users\dploy\.openclaw\workspace\basicshirts_web\master_clean")

wb = openpyxl.load_workbook(SRC, data_only=True, keep_vba=True)
ws_dd = wb["DropDown"]
ws_label = wb["Label"]

OUT_DIR.mkdir(parents=True, exist_ok=True)


def uniq_col(ws, col: int, start_row: int = 1):
    vals=[]
    for r in range(start_row, ws.max_row + 1):
        v = ws.cell(r, col).value
        if v is None:
            continue
        s = str(v).strip() if isinstance(v, str) else str(v)
        s = s.strip()
        if not s:
            continue
        vals.append(s)
    seen=set(); out=[]
    for x in vals:
        if x in seen:
            continue
        seen.add(x)
        out.append(x)
    return out


# DropDown columns: U(display packaging), V(transit package), W(label)
# Excel columns are 1-indexed: U=21, V=22, W=23
u_vals = uniq_col(ws_dd, 21, 1)
v_vals = uniq_col(ws_dd, 22, 1)
w_vals = uniq_col(ws_dd, 23, 1)

# write dropdowns
for fname, vals in [
    ("dropdown_display_packaging.csv", u_vals),
    ("dropdown_transit_package.csv", v_vals),
    ("dropdown_label.csv", w_vals),
]:
    with (OUT_DIR / fname).open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["value"])
        for x in vals:
            w.writerow([x])

# Label sheet lookup A->B for display packaging prices
rows=[]
for r in range(1, ws_label.max_row + 1):
    a = ws_label.cell(r, 1).value
    b = ws_label.cell(r, 2).value
    if a is None and b is None:
        continue
    if a is None or b is None:
        continue
    key = str(a).strip() if isinstance(a, str) else str(a)
    key = key.strip()
    if not key or key.lower() in ("display packaging", "item"):
        continue
    try:
        price = float(b)
    except Exception:
        continue
    rows.append((key, price))

seen=set(); uniq=[]
for k,p in rows:
    if k in seen:
        continue
    seen.add(k)
    uniq.append((k,p))

with (OUT_DIR / "label_display_packaging_price.csv").open("w", newline="", encoding="utf-8") as f:
    w=csv.writer(f)
    w.writerow(["display_packaging","price"])
    for k,p in uniq:
        w.writerow([k,p])

print("wrote dropdown_display_packaging.csv", len(u_vals))
print("wrote dropdown_transit_package.csv", len(v_vals))
print("wrote dropdown_label.csv", len(w_vals))
print("wrote label_display_packaging_price.csv", len(uniq))
