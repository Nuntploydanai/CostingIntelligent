import csv
from pathlib import Path
import openpyxl

SRC = r"C:\Users\dploy\Downloads\Design to Basic Shirt Tool First Trail Version 1 - 20250801-update latest.xlsm"
OUT = Path(r"C:\Users\dploy\.openclaw\workspace\basicshirts_web\master_clean\fabric_price_ar_as_lookup.csv")

wb = openpyxl.load_workbook(SRC, data_only=True, keep_vba=True)
ws = wb["Fabric Price"]

# Extract AR (44) -> AS (45)
COL_KEY = 44
COL_VAL = 45

rows = []
for r in range(1, ws.max_row + 1):
    k = ws.cell(r, COL_KEY).value
    v = ws.cell(r, COL_VAL).value
    if k is None or v is None:
        continue
    ks = str(k).strip()
    if not ks:
        continue
    try:
        vf = float(v)
    except Exception:
        continue
    rows.append((ks, vf))

# de-dup preserve order
seen=set(); uniq=[]
for k,v in rows:
    if k in seen: continue
    seen.add(k); uniq.append((k,v))

OUT.parent.mkdir(parents=True, exist_ok=True)
with OUT.open('w', newline='', encoding='utf-8') as f:
    w=csv.writer(f)
    w.writerow(['key','value'])
    for k,v in uniq:
        w.writerow([k,v])

print(f"wrote {OUT} rows={len(uniq)}")
