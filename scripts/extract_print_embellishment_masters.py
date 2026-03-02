import csv
from pathlib import Path
import openpyxl

SRC = r"C:\Users\dploy\Downloads\Design to Basic Shirt Tool First Trail Version 1 - 20250801-update latest.xlsm"
OUT_DIR = Path(r"C:\Users\dploy\.openclaw\workspace\basicshirts_web\master_clean")

wb = openpyxl.load_workbook(SRC, data_only=True, keep_vba=True)
ws = wb["Print_ and_Embroidery"]
OUT_DIR.mkdir(parents=True, exist_ok=True)

# Excel logic uses Print_ and_Embroidery sheet columns A (type) + B (dimension)
# for both dropdown behavior and default price lookup.

# Extract lookup for default price/each used by Data link N22:
# XLOOKUP(H22&J22, Print_ and_Embroidery!A:A&B:B, G:G)
# Extract columns A(1), B(2), G(7)
rows=[]
for r in range(1, ws.max_row+1):
    a=ws.cell(r,1).value
    b=ws.cell(r,2).value
    g=ws.cell(r,7).value
    if a is None and b is None and g is None:
        continue
    if a is None or b is None or g is None:
        continue
    a_s=str(a).strip() if isinstance(a,str) else str(a)
    b_s=str(b).strip() if isinstance(b,str) else str(b)
    if not a_s or not b_s:
        continue
    # skip headers
    if a_s.lower() in ("printing/embroidery", "printing", "embroidery") and b_s.lower() == "dimension":
        continue
    try:
        price=float(g)
    except Exception:
        continue
    rows.append((a_s,b_s,price))

seen=set(); uniq=[]
for a_s,b_s,price in rows:
    k=(a_s,b_s)
    if k in seen:
        continue
    seen.add(k)
    uniq.append((a_s,b_s,price))

with (OUT_DIR / 'print_embroidery_price_lookup.csv').open('w', newline='', encoding='utf-8') as f:
    w=csv.writer(f)
    w.writerow(['printing_embroidery','dimension','default_price_each'])
    for a_s,b_s,price in uniq:
        w.writerow([a_s,b_s,price])

# Dropdowns derived from the lookup pairs:
print_types=[]
dims_by_type={}
for a_s,b_s,_ in uniq:
    if a_s not in dims_by_type:
        dims_by_type[a_s]=[]
        print_types.append(a_s)
    if b_s not in dims_by_type[a_s]:
        dims_by_type[a_s].append(b_s)

with (OUT_DIR / 'dropdown_printing_embroidery.csv').open('w', newline='', encoding='utf-8') as f:
    w=csv.writer(f); w.writerow(['value']);
    for x in print_types:
        w.writerow([x])

# Keep this file for compatibility; UI now fetches filtered dimensions per type.
all_dims=[]
seen_dims=set()
for t in print_types:
    for d in dims_by_type[t]:
        if d in seen_dims: continue
        seen_dims.add(d)
        all_dims.append(d)
# Sort dimensions naturally (1X1 cm before 2X2 cm, etc.)
import re
def dim_sort_key(d):
    # Extract first number for sorting
    m = re.match(r'(\d+)', d)
    if m:
        return int(m.group(1))
    return 999
all_dims.sort(key=dim_sort_key)
with (OUT_DIR / 'dropdown_print_dimension.csv').open('w', newline='', encoding='utf-8') as f:
    w=csv.writer(f); w.writerow(['value']);
    for x in all_dims:
        w.writerow([x])

print('wrote dropdown_printing_embroidery.csv', len(print_types))
print('wrote dropdown_print_dimension.csv', len(all_dims))
print('wrote print_embroidery_price_lookup.csv', len(uniq))
