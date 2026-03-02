import csv
from pathlib import Path
import openpyxl

SRC = r"C:\Users\dploy\Downloads\Design to Basic Shirt Tool First Trail Version 1 - 20250801-update latest.xlsm"
OUT = Path(r"C:\Users\dploy\.openclaw\workspace\basicshirts_web\master_clean\fabric_price_lookup.csv")

wb = openpyxl.load_workbook(SRC, data_only=True, keep_vba=True)
ws = wb["Fabric Price"]

# Data link uses XLOOKUP(E7, Fabric Price!U:U, Fabric Price!W:W)
# So we extract columns U and W (21 and 23) for all non-empty keys.
U_COL = 21
W_COL = 23

rows = [["key", "value"]]
seen = set()
for r in range(1, ws.max_row + 1):
    k = ws.cell(r, U_COL).value
    v = ws.cell(r, W_COL).value
    if k is None or str(k).strip() == "":
        continue
    key = str(k).strip()
    if key in seen:
        continue
    seen.add(key)
    rows.append([key, v if v is not None else ""])

OUT.parent.mkdir(parents=True, exist_ok=True)
with OUT.open("w", newline="", encoding="utf-8") as f:
    w = csv.writer(f)
    w.writerows(rows)

print(f"wrote {OUT} rows={len(rows)-1}")
