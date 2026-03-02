import csv
from pathlib import Path
import openpyxl

SRC = r"C:\Users\dploy\Downloads\Design to Basic Shirt Tool First Trail Version 1 - 20250801-update latest.xlsm"
OUT = Path(r"C:\Users\dploy\.openclaw\workspace\basicshirts_web\master_clean\dropdown_using_part.csv")

wb = openpyxl.load_workbook(SRC, data_only=True, keep_vba=True)
ws = wb["Fabric Usage"]

# Data link uses XLOOKUP(K7&L7, Fabric Usage!C:C & Fabric Usage!D:D, ...)
# So Using Part should be in column D (heuristic based on that formula).
# We'll extract unique non-empty strings from column D.
COL = 4  # D

vals = []
for r in range(1, ws.max_row + 1):
    v = ws.cell(r, COL).value
    if v is None:
        continue
    if isinstance(v, str):
        s = v.strip()
        if not s:
            continue
        vals.append(s)
    else:
        vals.append(str(v))

# de-dup preserve order
seen = set()
uniq = []
for x in vals:
    if x not in seen:
        seen.add(x)
        uniq.append(x)

OUT.parent.mkdir(parents=True, exist_ok=True)
with OUT.open("w", newline="", encoding="utf-8") as f:
    w = csv.writer(f)
    w.writerow(["value"])
    for x in uniq:
        w.writerow([x])

print(f"wrote {OUT} count={len(uniq)}")
