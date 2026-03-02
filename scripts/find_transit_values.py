import openpyxl

SRC = r"C:\Users\dploy\Downloads\Design to Basic Shirt Tool First Trail Version 1 - 20250801-update latest.xlsm"

wb = openpyxl.load_workbook(SRC, data_only=True, keep_vba=True)
ws_dd = wb["DropDown"]

print("Searching for values like 36, 48 in all columns...")
for col in range(1, 30):
    for row in range(1, 100):
        v = ws_dd.cell(row, col).value
        if v in [36, 48, '36', '48']:
            print(f"  Found {v} at row {row}, col {col}")

print("\n\nLet me check what column has transit package values...")
print("Checking columns around V (22):")
for col in [20, 21, 22, 23, 24, 25]:
    print(f"\nColumn {col}:")
    for row in range(1, 20):
        v = ws_dd.cell(row, col).value
        if v is not None:
            print(f"  Row {row}: {v}")
