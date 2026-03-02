import csv
from pathlib import Path
import openpyxl

SRC = r"C:\Users\dploy\Downloads\Design to Basic Shirt Tool First Trail Version 1 - 20250801-update latest.xlsm"
OUT = Path(r"C:\Users\dploy\.openclaw\workspace\basicshirts_web\master_clean\fabric_width_condition_map.csv")

wb = openpyxl.load_workbook(SRC, data_only=True, keep_vba=True)
dl = wb["Data link"]

# Data link I38: XLOOKUP(E26&E27, B38:B51 & E38:E51, F38:F51)
rows=[]
for r in range(38,52):
    sil = dl.cell(r,2).value  # B
    seam = dl.cell(r,5).value # E
    width = dl.cell(r,6).value # F
    if sil is None and seam is None and width is None:
        continue
    sil_s = (str(sil).strip() if sil is not None else "")
    seam_s = (str(seam).strip() if seam is not None else "")
    if width is None or sil_s=="" or seam_s=="":
        continue
    try:
        width_f = float(width)
    except Exception:
        continue
    rows.append((sil_s,seam_s,width_f))

OUT.parent.mkdir(parents=True, exist_ok=True)
with OUT.open('w', newline='', encoding='utf-8') as f:
    w=csv.writer(f)
    w.writerow(['silhouette','seam','width_in'])
    for a,b,c in rows:
        w.writerow([a,b,c])

print(f"wrote {OUT} rows={len(rows)}")
