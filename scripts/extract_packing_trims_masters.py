import csv
from pathlib import Path
import openpyxl

SRC = r"C:\Users\dploy\Downloads\Design to Basic Shirt Tool First Trail Version 1 - 20250801-update latest.xlsm"
OUT_DIR = Path(r"C:\Users\dploy\.openclaw\workspace\basicshirts_web\master_clean")

wb = openpyxl.load_workbook(SRC, data_only=True, keep_vba=True)
ws = wb["Packing&Trims"]

OUT_DIR.mkdir(parents=True, exist_ok=True)

# Trims table: B=Trims, C=Price/unit, D=Unit
item_rows = []
for r in range(2, ws.max_row + 1):
    b = ws.cell(r, 2).value
    c = ws.cell(r, 3).value
    d = ws.cell(r, 4).value
    if b is None and c is None and d is None:
        continue
    if b is None:
        continue
    item = str(b).strip()
    if not item:
        continue
    # price might be blank for some items; keep None
    price = None
    if c not in (None, ""):
        try:
            price = float(c)
        except Exception:
            price = None
    unit = str(d).strip() if d is not None else ""
    item_rows.append((item, price, unit))

# de-dup items preserve first
seen=set(); uniq_items=[]
for item, price, unit in item_rows:
    if item in seen:
        continue
    seen.add(item)
    uniq_items.append((item, price, unit))

# write dropdown trims type
with (OUT_DIR / "dropdown_trims_type.csv").open('w', newline='', encoding='utf-8') as f:
    w=csv.writer(f)
    w.writerow(['value'])
    for item, _, _ in uniq_items:
        w.writerow([item])

# write lookup price + unit
with (OUT_DIR / "packing_trims_item_price_unit.csv").open('w', newline='', encoding='utf-8') as f:
    w=csv.writer(f)
    w.writerow(['trims_type','default_price_each','unit'])
    for item, price, unit in uniq_items:
        w.writerow([item, '' if price is None else price, unit])

# Usage table: K=Trims, L=Using Part, M=SMALL, N=LARGE
usage_rows=[]
for r in range(2, ws.max_row + 1):
    k = ws.cell(r, 11).value
    l = ws.cell(r, 12).value
    m = ws.cell(r, 13).value
    n = ws.cell(r, 14).value
    if k is None and l is None and m is None and n is None:
        continue
    if k is None or l is None:
        continue
    trims = str(k).strip(); part = str(l).strip()
    if not trims or not part:
        continue
    small = None
    if m not in (None, ""):
        try:
            small = float(m)
        except Exception:
            small = None
    large = None
    if n not in (None, ""):
        try:
            large = float(n)
        except Exception:
            large = None
    if small is None and large is None:
        continue
    usage_rows.append((trims, part, small, large))

# de-dup by (trims, part) keep first
seen=set(); uniq_usage=[]
for trims, part, small, large in usage_rows:
    key=(trims, part)
    if key in seen:
        continue
    seen.add(key)
    uniq_usage.append((trims, part, small, large))

with (OUT_DIR / "packing_trims_usage.csv").open('w', newline='', encoding='utf-8') as f:
    w=csv.writer(f)
    w.writerow(['trims_type','garment_part','usage_small','usage_large'])
    for trims, part, small, large in uniq_usage:
        w.writerow([trims, part, '' if small is None else small, '' if large is None else large])

# dropdown garment part (unique L)
seen=set(); parts=[]
for _, part, _, _ in uniq_usage:
    if part in seen:
        continue
    seen.add(part)
    parts.append(part)
with (OUT_DIR / "dropdown_garment_part_trim.csv").open('w', newline='', encoding='utf-8') as f:
    w=csv.writer(f)
    w.writerow(['value'])
    for p in parts:
        w.writerow([p])

print('wrote:',
      'dropdown_trims_type.csv', len(uniq_items),
      'packing_trims_item_price_unit.csv', len(uniq_items),
      'packing_trims_usage.csv', len(uniq_usage),
      'dropdown_garment_part_trim.csv', len(parts))
