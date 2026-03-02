import openpyxl

SRC = r"C:\Users\dploy\Downloads\Design to Basic Shirt Tool First Trail Version 1 - 20250801-update latest.xlsm"

wb = openpyxl.load_workbook(SRC, data_only=True, keep_vba=True)
ws_dd = wb["DropDown"]

print("Checking DropDown sheet columns U, V, W:")
print("\nColumn U (21) - first 10 rows:")
for r in range(1, 11):
    v = ws_dd.cell(r, 21).value
    print(f"  Row {r}: {v}")

print("\nColumn V (22) - first 30 rows:")
for r in range(1, 31):
    v = ws_dd.cell(r, 22).value
    print(f"  Row {r}: {v}")

print("\nColumn W (23) - first 10 rows:")
for r in range(1, 11):
    v = ws_dd.cell(r, 23).value
    print(f"  Row {r}: {v}")
