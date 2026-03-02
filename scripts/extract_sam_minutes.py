"""
Extract SAM (minutes) lookup table from SAM&Product EFF% sheet
"""
import openpyxl
import csv
from pathlib import Path

SRC = r"C:\Users\dploy\Downloads\Design to Basic Shirt Tool First Trail Version 1 - 20250801-update latest.xlsm"
OUT_DIR = Path(r"C:\Users\dploy\.openclaw\workspace\basicshirts_web\master_clean")

wb = openpyxl.load_workbook(SRC, data_only=True, keep_vba=True)
ws_sam = wb['SAM&Product EFF%']

OUT_DIR.mkdir(parents=True, exist_ok=True)

print("=" * 80)
print("EXTRACTING SAM (MINUTES) LOOKUP TABLE")
print("=" * 80)

sam_data = []

for row in range(2, 100):
    gender = ws_sam.cell(row, 1).value  # A column
    product = ws_sam.cell(row, 2).value  # B column
    seam = ws_sam.cell(row, 3).value  # C column
    size = ws_sam.cell(row, 4).value  # D column
    sam_minutes = ws_sam.cell(row, 5).value  # E column

    if gender and product and seam and size and sam_minutes:
        sam_data.append({
            'gender': str(gender).strip(),
            'product': str(product).strip(),
            'seam': str(seam).strip(),
            'size': str(size).strip(),
            'sam_minutes': float(sam_minutes)
        })

print(f"\nFound {len(sam_data)} rows of SAM data")

# Write to CSV
with (OUT_DIR / "sam_minutes_lookup.csv").open("w", newline="", encoding="utf-8") as f:
    w = csv.DictWriter(f, fieldnames=['gender', 'product', 'seam', 'size', 'sam_minutes'])
    w.writeheader()
    for row in sam_data:
        w.writerow(row)

print(f"\nWrote sam_minutes_lookup.csv")

# Verify Men + T-Shirt (Crew Neck) + No seam + S-XL
print("\n" + "=" * 80)
print("VERIFICATION: Men + T-Shirt (Crew Neck) + No seam + S-XL")
print("=" * 80)
for row in sam_data:
    if row['gender'] == 'Men' and 'Crew Neck' in row['product'] and row['seam'] == 'No seam' and row['size'] == 'S-XL':
        print(f"Found: {row}")
