import csv
from pathlib import Path
import openpyxl

SRC = r"C:\Users\dploy\Downloads\Design to Basic Shirt Tool First Trail Version 1 - 20250801-update latest.xlsm"
OUT_DIR = Path(r"C:\Users\dploy\.openclaw\workspace\basicshirts_web\master_clean")
OUT_DIR.mkdir(parents=True, exist_ok=True)

wb = openpyxl.load_workbook(SRC, data_only=True, keep_vba=True)
ws = wb["Fabric Price"]

# Heuristic: look for a header row containing both 'Fabric' and 'Type'
header_row = None
col_map = {}
for r in range(1, min(20, ws.max_row) + 1):
    row_vals = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(r, c).value
        if isinstance(v, str) and v.strip():
            row_vals[v.strip().lower()] = c
    # detect likely columns
    fabric_col = None
    type_col = None
    for k, c in row_vals.items():
        if k == "fabric" or "content" in k:
            fabric_col = c
        if k == "type":
            type_col = c
    if fabric_col and type_col:
        header_row = r
        col_map = {"fabric": fabric_col, "type": type_col}
        break

# Fallback to known structure from extracted master CSV sample:
# Item,Fabric,Type,Weight,Cut width,Price/Yd
if header_row is None:
    header_row = 1
    # try to match exact headers in row 1
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if isinstance(v, str) and v.strip():
            headers[v.strip().lower()] = c
    col_map["fabric"] = headers.get("fabric")
    col_map["type"] = headers.get("type")

fabric_col = col_map.get("fabric")
type_col = col_map.get("type")

if not fabric_col or not type_col:
    raise SystemExit(f"Could not detect Fabric/Type columns in Fabric Price sheet. Detected: {col_map}")

fabric_values = []
type_values = []

for r in range(header_row + 1, ws.max_row + 1):
    fv = ws.cell(r, fabric_col).value
    tv = ws.cell(r, type_col).value
    if isinstance(fv, str):
        s = fv.strip()
        if s:
            fabric_values.append(s)
    if isinstance(tv, str):
        s = tv.strip()
        if s:
            type_values.append(s)

# de-dup preserve order

def uniq(xs):
    seen = set(); out=[]
    for x in xs:
        if x not in seen:
            seen.add(x); out.append(x)
    return out

fabric_values = uniq(fabric_values)
type_values = uniq(type_values)

# write CSVs

def write_list(path: Path, values: list[str]):
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["value"])
        for v in values:
            w.writerow([v])

# Remove any accidental type values from contents
fabric_types_set = set(type_values)
fabric_values = [v for v in fabric_values if v not in fabric_types_set]

write_list(OUT_DIR / "dropdown_fabric_contents.csv", fabric_values)
# NOTE: dropdown_fabric_type.csv is built from Conversion (more complete). Keep this file too for reference.
write_list(OUT_DIR / "_dropdown_fabric_type_from_fabric_price.csv", type_values)

print("wrote fabric dropdowns contents=", len(fabric_values), "types(from fabric price)=", len(type_values))
