import openpyxl

SRC = r"C:\Users\dploy\Downloads\Design to Basic Shirt Tool First Trail Version 1 - 20250801-update latest.xlsm"

wb = openpyxl.load_workbook(SRC, data_only=True, keep_vba=True)
ws_dd = wb["DropDown"]

print("Full list of values in Column V (22) - Transit Package:")
vals = []
for row in range(1, ws_dd.max_row + 1):
    v = ws_dd.cell(row, 22).value
    if v is not None:
        s = str(v).strip()
        if s and s != "TRANSIT PACKAGE":
            vals.append(s)

print(f"Total values: {len(vals)}")
print(f"First 50: {vals[:50]}")
print(f"\nLast 20: {vals[-20:]}")

# Check if 36, 48 are in there
if "36" in vals:
    print(f"\n✓ 36 found at position {vals.index('36')}")
if "48" in vals:
    print(f"✓ 48 found at position {vals.index('48')}")
