import csv
from pathlib import Path
import openpyxl

SRC = r"C:\Users\dploy\Downloads\Design to Basic Shirt Tool First Trail Version 1 - 20250801-update latest.xlsm"
OUT = Path(r"C:\Users\dploy\.openclaw\workspace\basicshirts_web\master_clean\product_part_key_map.csv")

wb = openpyxl.load_workbook(SRC, data_only=True, keep_vba=True)
ws = wb["Data link"]

# Data link K7 mapping table: B38:C51 -> D38:D51
rows = []
for r in range(38, 52):
    sil = ws.cell(r, 2).value  # B
    seam = ws.cell(r, 3).value  # C
    key = ws.cell(r, 4).value  # D
    if sil is None and seam is None and key is None:
        continue
    sil_s = (str(sil).strip() if sil is not None else "")
    seam_s = (str(seam).strip() if seam is not None else "")
    key_s = (str(key).strip() if key is not None else "")
    if not (sil_s and seam_s and key_s):
        continue
    rows.append((sil_s, seam_s, key_s))

OUT.parent.mkdir(parents=True, exist_ok=True)
with OUT.open("w", newline="", encoding="utf-8") as f:
    w = csv.writer(f)
    w.writerow(["silhouette", "seam", "k7_key"])
    for sil_s, seam_s, key_s in rows:
        w.writerow([sil_s, seam_s, key_s])

print(f"wrote {OUT} rows={len(rows)}")
