"""
Extract Step 6: Manufacturing Cost master data
"""
import csv
from pathlib import Path
import openpyxl

SRC = r"C:\Users\dploy\Downloads\Design to Basic Shirt Tool First Trail Version 1 - 20250801-update latest.xlsm"
OUT_DIR = Path(r"C:\Users\dploy\.openclaw\workspace\basicshirts_web\master_clean")

wb = openpyxl.load_workbook(SRC, data_only=True, keep_vba=True)
ws_dl = wb['Data link']
ws_cr = wb[' Cost Rate']

OUT_DIR.mkdir(parents=True, exist_ok=True)

# Extract manufacturing cost data for each country (rows 39-44)
print("=== Extracting Manufacturing Cost Data ===")
manufacturing_data = []
for row in range(39, 45):
    country = ws_dl.cell(row, 12).value  # L column - country
    minutes = ws_dl.cell(row, 17).value  # Q column - minutes
    cost_rate = ws_dl.cell(row, 19).value  # S column - cost rate
    efficiency = ws_dl.cell(row, 21).value  # U column - efficiency
    total_cost = ws_dl.cell(row, 26).value  # Z column - total cost

    print(f"Row {row}: {country}, {minutes}, {cost_rate}, {efficiency}, {total_cost}")

    if country:
        manufacturing_data.append({
            'country': country,
            'minutes': minutes,
            'cost_rate': cost_rate,
            'efficiency': efficiency,
            'total_cost': total_cost
        })

# Write to CSV
with (OUT_DIR / "manufacturing_cost_by_country.csv").open("w", newline="", encoding="utf-8") as f:
    w = csv.DictWriter(f, fieldnames=['country', 'minutes', 'cost_rate', 'efficiency', 'total_cost'])
    w.writeheader()
    for row in manufacturing_data:
        w.writerow(row)

print(f"\nWrote manufacturing_cost_by_country.csv with {len(manufacturing_data)} countries")

# Extract overhead costs (AC column)
print("\n=== Extracting Overhead Costs ===")
ws_bs = wb['Basic Shirts Costing Tool']
overhead_costs = {}
for row in range(2, 14):
    label = ws_bs.cell(row, 27).value  # AA column - label
    value = ws_bs.cell(row, 29).value  # AC column - value
    if label and value:
        overhead_costs[label] = value
        print(f"  {label}: {value}")

with (OUT_DIR / "overhead_costs.csv").open("w", newline="", encoding="utf-8") as f:
    w = csv.writer(f)
    w.writerow(['label', 'value'])
    for label, value in overhead_costs.items():
        w.writerow([label, value])

print(f"\nWrote overhead_costs.csv with {len(overhead_costs)} items")
