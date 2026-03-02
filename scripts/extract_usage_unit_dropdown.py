import csv
from pathlib import Path
import openpyxl

SRC = r"C:\Users\dploy\Downloads\Design to Basic Shirt Tool First Trail Version 1 - 20250801-update latest.xlsm"
OUT_DIR = Path(r"C:\Users\dploy\.openclaw\workspace\basicshirts_web\master_clean")

wb = openpyxl.load_workbook(SRC, data_only=True, keep_vba=True)
ws_dd = wb["DropDown"]

OUT_DIR.mkdir(parents=True, exist_ok=True)

# Extract column L (12) - Usage/Unit dropdown (X1, X2, ... X30)
# But we'll output just the numbers (1, 2, 3... 30)
vals=[]
for r in range(1, ws_dd.max_row + 1):
    v = ws_dd.cell(r, 12).value  # column L
    if v is None:
        continue
    s = str(v).strip() if isinstance(v, str) else str(v)
    s = s.strip()
    if not s:
        continue
    # skip header-like values
    if s.lower() in ("usage/unit", "usage", "unit", "packing_no"):
        continue
    # Convert X13 -> 13, or keep as-is if just a number
    if s.upper().startswith("X") and len(s) > 1:
        s = s[1:]  # Remove X prefix
    vals.append(s)

seen=set(); out=[]
for x in vals:
    if x in seen:
        continue
    seen.add(x)
    out.append(x)

# Sort numerically
def num_sort_key(x):
    try:
        return int(x)
    except Exception:
        return 999
out.sort(key=num_sort_key)

with (OUT_DIR / "dropdown_usage_unit.csv").open("w", newline="", encoding="utf-8") as f:
    w=csv.writer(f)
    w.writerow(["value"])
    for x in out:
        w.writerow([x])

print("wrote dropdown_usage_unit.csv", len(out))
print("values:", out[:10], "..." if len(out)>10 else "")
