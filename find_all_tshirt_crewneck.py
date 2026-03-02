"""
Find ALL rows with T-Shirt (Crew Neck) + No seam to see which one is correct
"""
import openpyxl

SRC = r"C:\Users\dploy\Downloads\Design to Basic Shirt Tool First Trail Version 1 - 20250801-update latest.xlsm"

wb = openpyxl.load_workbook(SRC, data_only=True, keep_vba=True)
ws_sam = wb['SAM&Product EFF%']

print("=" * 80)
print("FIND ALL T-Shirt (Crew Neck) + No seam ROWS")
print("=" * 80)

for row_num in range(1, 100):
    a_val = ws_sam.cell(row_num, 1).value  # Gender
    b_val = ws_sam.cell(row_num, 2).value  # Product shape
    c_val = ws_sam.cell(row_num, 3).value  # Side Seam
    d_val = ws_sam.cell(row_num, 4).value  # Size
    e_val = ws_sam.cell(row_num, 5).value  # SAM

    if b_val and 'T-Shirt (Crew Neck)' in str(b_val):
        if c_val and 'No seam' in str(c_val):
            print(f"\nRow {row_num}:")
            print(f"  A{row_num} (Gender): {a_val}")
            print(f"  B{row_num} (Product): {b_val}")
            print(f"  C{row_num} (Seam): {c_val}")
            print(f"  D{row_num} (Size): {d_val}")
            print(f"  E{row_num} (SAM): {e_val}")

            # Also check what's in F and G
            f_val = ws_sam.cell(row_num, 6).value
            g_val = ws_sam.cell(row_num, 7).value
            print(f"  F{row_num} (DEFAULT?): {f_val}")
            print(f"  G{row_num} (EFF%?): {g_val}")

print("\n" + "=" * 80)
print("Which row has Men + S-XL with SAM = 2.3?")
print("=" * 80)
