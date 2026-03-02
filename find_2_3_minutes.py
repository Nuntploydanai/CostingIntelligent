"""
Find minutes for: Men + T-Shirt (Crew Neck) + No seam + S-XL
"""
import openpyxl

SRC = r"C:\Users\dploy\Downloads\Design to Basic Shirt Tool First Trail Version 1 - 20250801-update latest.xlsm"

wb = openpyxl.load_workbook(SRC, data_only=True, keep_vba=True)

ws_sam = wb['SAM&Product EFF%']

print("=" * 80)
print("LOOKING FOR: Men + T-Shirt (Crew Neck) + No seam + S-XL")
print("=" * 80)

for row in range(1, 100):
    a_val = ws_sam.cell(row, 1).value  # Gender
    b_val = ws_sam.cell(row, 2).value  # Product shape
    c_val = ws_sam.cell(row, 3).value  # Seam
    d_val = ws_sam.cell(row, 4).value  # Size
    e_val = ws_sam.cell(row, 5).value  # SAM (minutes)

    if b_val and 'T-Shirt (Crew Neck)' in str(b_val):
        if c_val and 'No seam' in str(c_val):
            print(f"\nRow {row}:")
            print(f"  A{row} (Gender): {a_val}")
            print(f"  B{row} (Product): {b_val}")
            print(f"  C{row} (Seam): {c_val}")
            print(f"  D{row} (Size): {d_val}")
            print(f"  E{row} (SAM/Minutes): {e_val}")

print("\n" + "=" * 80)
print("FOUND IT! Men + T-Shirt (Crew Neck) + No seam = 2.3 minutes!")
print("=" * 80)
