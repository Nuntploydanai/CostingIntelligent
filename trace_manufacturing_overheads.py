"""
Trace complete manufacturing cost including overheads
"""
import openpyxl

SRC = r"C:\Users\dploy\Downloads\Design to Basic Shirt Tool First Trail Version 1 - 20250801-update latest.xlsm"

wb = openpyxl.load_workbook(SRC, data_only=True, keep_vba=True)
ws_dl = wb['Data link']

print("=" * 80)
print("COMPLETE MANUFACTURING COST CHAIN (WITH OVERHEADS)")
print("=" * 80)

print("\n=== DATA LINK Row 30 (INDIA) - All columns ===")
cols_to_check = ['J', 'K', 'L', 'N', 'Q', 'R', 'V', 'W', 'X', 'Z']
for col in cols_to_check:
    val = ws_dl[f'{col}30'].value
    print(f"{col}30: {val}")

print("\n=== Formulas for V, W, X, Z ===")
wb2 = openpyxl.load_workbook(SRC, data_only=False, keep_vba=True)
ws_dl2 = wb2['Data link']

for col in ['V', 'W', 'X', 'Z']:
    formula = ws_dl2[f'{col}30'].value
    print(f"{col}30 formula: {formula}")

print("\n=== OVERHEAD COSTS (AC column in Basic Shirts) ===")
ws_bs = wb['Basic Shirts Costing Tool']
for row in range(1, 15):
    label = ws_bs.cell(row, 27).value  # AA column
    value = ws_bs.cell(row, 29).value  # AC column
    if label and value is not None:
        print(f"AC{row} ({label}): {value}")

print("\n" + "=" * 80)
print("KEY FINDING:")
print("Z30 = 2.29981 (includes overheads)")
print("R30 = basic cost only")
print("=" * 80)
