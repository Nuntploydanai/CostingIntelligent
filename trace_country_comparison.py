"""
Trace Excel cost comparison for different countries
"""
import openpyxl

SRC = r"C:\Users\dploy\OneDrive\Documents\Design to Basic Shirt Tool First Trail Version 1 - 20250801-update latest.xlsm"

wb = openpyxl.load_workbook(SRC, data_only=False, keep_vba=True)

print("=" * 80)
print("TRACING COST COMPARISON IN EXCEL")
print("=" * 80)

# Check Manufacturing sheet for multiple countries
if 'Manufacturing' in wb.sheetnames:
    ws_mfg = wb['Manufacturing']
    print("\nManufacturing Sheet:")
    for row in range(1, 30):
        row_data = []
        for col in range(1, 20):
            val = ws_mfg.cell(row, col).value
            if val is not None:
                row_data.append(str(val)[:20])
        if row_data:
            print(f"  Row {row}: {' | '.join(row_data)}")

# Check Basic Shirts Costing Tool for country comparison
if 'Basic Shirts Costing Tool' in wb.sheetnames:
    ws_bs = wb['Basic Shirts Costing Tool']

    print("\n" + "=" * 80)
    print("BASIC SHIRTS COSTING TOOL - COUNTRY COMPARISON")
    print("=" * 80)

    # Look for country comparison section
    for row in range(1, 50):
        for col in range(1, 40):
            val = ws_bs.cell(row, col).value
            if val and isinstance(val, str):
                if any(country in val.upper() for country in ['INDIA', 'BANGLADESH', 'CHINA', 'VIETNAM', 'INDONESIA', 'CAMBODIA']):
                    col_letter = openpyxl.utils.get_column_letter(col)
                    print(f"\n{col_letter}{row}: {val}")

                    # Check next few cells
                    for next_col in range(col, col + 15):
                        next_val = ws_bs.cell(row, next_col).value
                        if next_val is not None:
                            next_col_letter = openpyxl.utils.get_column_letter(next_col)
                            print(f"  {next_col_letter}{row}: {next_val}")

# Check for comparison table or formula
print("\n" + "=" * 80)
print("LOOKING FOR COST COMPARISON FORMULA")
print("=" * 80)

# Check if there's a table that calculates FOB for each country
ws_bs = wb['Basic Shirts Costing Tool']
for row in range(1, 100):
    for col in range(1, 50):
        val = ws_bs.cell(row, col).value
        if val and isinstance(val, str) and 'FOB' in val.upper():
            col_letter = openpyxl.utils.get_column_letter(col)
            print(f"\nFound FOB at {col_letter}{row}: {val}")

            # Show context
            for r in range(max(1, row-2), min(100, row+10)):
                row_data = []
                for c in range(max(1, col-2), min(50, col+10)):
                    cell_val = ws_bs.cell(r, c).value
                    if cell_val is not None:
                        row_data.append(f"{openpyxl.utils.get_column_letter(c)}{r}:{str(cell_val)[:15]}")
                if row_data:
                    print(f"  {' | '.join(row_data)}")

print("\n" + "=" * 80)
print("MANUFACTURING COST STRUCTURE")
print("=" * 80)

# Check how manufacturing cost varies by country
from pathlib import Path
import csv

cost_rate_path = Path(r"C:\Users\dploy\.openclaw\workspace\basicshirts_web\master_clean\cost_rate.csv")
print("\nCost Rate by Country:")
with open(cost_rate_path, 'r', encoding='utf-8') as f:
    reader = csv.DictReader(f)
    for row in reader:
        print(f"  {row.get('country', 'N/A'):15} - Rate: {row.get('cost_rate', 'N/A')}")
