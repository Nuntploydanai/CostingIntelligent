"""
Check if Data link should pull from D13-D16 instead of D16-D20
"""
import openpyxl

SRC = r"C:\Users\dploy\Downloads\Design to Basic Shirt Tool First Trail Version 1 - 20250801-update latest.xlsm"

wb = openpyxl.load_workbook(SRC, data_only=True, keep_vba=True)

ws_bs = wb['Basic Shirts Costing Tool']
ws_dl = wb['Data link']

print("=" * 80)
print("BASIC SHIRTS - Check both D13-D16 AND D16-D20")
print("=" * 80)

print("\n=== D13-D16 ===")
print(f"D13: {ws_bs['D13'].value}")
print(f"D14: {ws_bs['D14'].value}")
print(f"D15: {ws_bs['D15'].value}")
print(f"D16: {ws_bs['D16'].value}")

print("\n=== D16-D20 ===")
print(f"D16: {ws_bs['D16'].value}")
print(f"D17: {ws_bs['D17'].value}")
print(f"D18: {ws_bs['D18'].value}")
print(f"D19: {ws_bs['D19'].value}")
print(f"D20: {ws_bs['D20'].value}")

print("\n" + "=" * 80)
print("DATA LINK current values (from D16-D20):")
print("=" * 80)
print(f"E25: {ws_dl['E25'].value}")
print(f"E26: {ws_dl['E26'].value}")
print(f"E27: {ws_dl['E27'].value}")
print(f"E29: {ws_dl['E29'].value}")
print(f"J30 (minutes): {ws_dl['J30'].value}")

print("\n" + "=" * 80)
print("HYPOTHESIS: Data link should pull from D13-D16")
print("If so, it would be:")
print(f"  Gender: {ws_bs['D13'].value}")
print(f"  Silhouette: {ws_bs['D14'].value}")
print(f"  Seam: {ws_bs['D15'].value}")
print(f"  Size: {ws_bs['D16'].value}")
print("=" * 80)

# Manually test lookup with D13-D16
ws_sam = wb['SAM&Product EFF%']

gender = ws_bs['D13'].value
silhouette = ws_bs['D14'].value
seam = ws_bs['D15'].value
size = ws_bs['D16'].value

print(f"\nLooking up: {gender} + {silhouette} + {seam} + {size}")

# Search for matching row
for row_num in range(2, 100):
    a_val = ws_sam.cell(row_num, 1).value
    b_val = ws_sam.cell(row_num, 2).value
    c_val = ws_sam.cell(row_num, 3).value
    d_val = ws_sam.cell(row_num, 4).value
    e_val = ws_sam.cell(row_num, 5).value

    if a_val == gender and b_val == silhouette and c_val == seam and d_val == size:
        print(f"\nFOUND at row {row_num}:")
        print(f"  Gender: {a_val}")
        print(f"  Product: {b_val}")
        print(f"  Seam: {c_val}")
        print(f"  Size: {d_val}")
        print(f"  SAM (minutes): {e_val}")
        break
