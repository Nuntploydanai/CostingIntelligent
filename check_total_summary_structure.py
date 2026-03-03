"""
Check Total Cost Summary structure from Excel
"""
import openpyxl

SRC = r"C:\Users\dploy\OneDrive\Documents\Design to Basic Shirt Tool First Trail Version 1 - 20250801-update latest.xlsm"

wb = openpyxl.load_workbook(SRC, data_only=True, keep_vba=True)
ws_bs = wb['Basic Shirts Costing Tool']

print("=" * 80)
print("TOTAL COST SUMMARY STRUCTURE FROM EXCEL")
print("=" * 80)

# Look for the Total Cost Summary section
# It's usually at the bottom of the Basic Shirts Costing Tool sheet

print("\nSearching for 'Total Cost Summary' or similar labels...")
print("\nColumn A labels (rows 70-100):")
for row in range(70, 100):
    a_val = ws_bs.cell(row, 1).value
    if a_val and any(keyword in str(a_val).lower() for keyword in ['total', 'cost', 'summary', 'price', 'fob']):
        print(f"Row {row}: A={a_val}")
        # Also check column B, C, D for values
        b_val = ws_bs.cell(row, 2).value
        c_val = ws_bs.cell(row, 3).value
        d_val = ws_bs.cell(row, 4).value
        k_val = ws_bs.cell(row, 11).value  # Column K (total)
        print(f"         B={b_val}, C={c_val}, D={d_val}, K={k_val}")
        print()

print("\n" + "=" * 80)
print("Looking for specific sections...")
print("=" * 80)

# Check if there are overhead costs
print("\nSearching for overhead costs...")
for row in range(1, 150):
    a_val = ws_bs.cell(row, 1).value
    if a_val and 'overhead' in str(a_val).lower():
        print(f"Row {row}: {a_val}")
        k_val = ws_bs.cell(row, 11).value
        print(f"  Total (K): {k_val}")

# Check for FOB price
print("\nSearching for FOB price...")
for row in range(1, 150):
    a_val = ws_bs.cell(row, 1).value
    if a_val and 'fob' in str(a_val).lower():
        print(f"Row {row}: {a_val}")
        k_val = ws_bs.cell(row, 11).value
        print(f"  Total (K): {k_val}")

# Check for profit margin
print("\nSearching for profit margin...")
for row in range(1, 150):
    a_val = ws_bs.cell(row, 1).value
    if a_val and 'profit' in str(a_val).lower():
        print(f"Row {row}: {a_val}")
        k_val = ws_bs.cell(row, 11).value
        print(f"  Total (K): {k_val}")

print("\n" + "=" * 80)
print("Checking Data link sheet for total cost summary...")
print("=" * 80)

ws_dl = wb['Data link']

print("\nData link - looking for total cost formulas...")
for row in range(1, 50):
    for col in range(1, 30):
        val = ws_dl.cell(row, col).value
        if val and isinstance(val, str) and 'total' in val.lower():
            col_letter = openpyxl.utils.get_column_letter(col)
            print(f"{col_letter}{row}: {val}")
