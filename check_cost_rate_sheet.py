"""
Check for Cost Rate sheet with all countries
"""
import openpyxl

SRC = r"C:\Users\dploy\OneDrive\Documents\Design to Basic Shirt Tool First Trail Version 1 - 20250801-update latest.xlsm"

wb = openpyxl.load_workbook(SRC, data_only=True, keep_vba=True)

print("=" * 80)
print("CHECKING FOR COST RATE SHEET")
print("=" * 80)

print("\nAvailable sheets:")
for sheet_name in wb.sheetnames:
    if 'cost' in sheet_name.lower() or 'rate' in sheet_name.lower():
        print(f"  - {sheet_name}")

# Check if "Cost Rate" sheet exists
if 'Cost Rate' in wb.sheetnames:
    ws_cr = wb['Cost Rate']

    print("\n" + "=" * 80)
    print("COST RATE SHEET FOUND")
    print("=" * 80)

    # Print first 10 rows
    for row in range(1, 20):
        a_val = ws_cr.cell(row, 1).value  # Country
        o_val = ws_cr.cell(row, 15).value  # Column O (cost rate?)

        if a_val or o_val:
            print(f"Row {row}: A={a_val}, O={o_val}")
else:
    print("\nCost Rate sheet NOT found")
    print("\nChecking Data link for cost rate formulas...")

    ws_dl = wb['Data link']

    # Check L30 formula
    wb_formulas = openpyxl.load_workbook(SRC, data_only=False, keep_vba=True)
    ws_dl_f = wb_formulas['Data link']

    print(f"\nL30 formula: {ws_dl_f['L30'].value}")
    print(f"L30 value: {wb['Data link']['L30'].value}")
