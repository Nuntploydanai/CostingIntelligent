"""
Extract SAM (minutes) lookup table from user's Excel file
"""
import openpyxl
import csv
from pathlib import Path

SRC = r"C:\Users\dploy\OneDrive\Documents\Design to Basic Shirt Tool First Trail Version 1 - 20250801-update latest.xlsm"
OUT_DIR = Path(r"C:\Users\dploy\.openclaw\workspace\basicshirts_web\master_clean")

wb = openpyxl.load_workbook(SRC, data_only=True, keep_vba=True)
ws_sam = wb['SAM&Product EFF%']

OUT_DIR.mkdir(parents=True, exist_ok=True)

print("=" * 80)
print("EXTRACTING SAM DATA FROM USER'S EXCEL FILE")
print("=" * 80)

sam_data = []

for row in range(2, 200):
    a_val = ws_sam.cell(row, 1).value  # Gender
    b_val = ws_sam.cell(row, 2).value  # Product shape
    c_val = ws_sam.cell(row, 3).value  # Seam
    d_val = ws_sam.cell(row, 4).value  # Size
    e_val = ws_sam.cell(row, 5).value  # SAM (minutes)

    if a_val and b_val and c_val and d_val and e_val is not None:
        # Fix seam casing to match dropdown
        seam = str(c_val).strip()
        if seam == "No seam":
            seam = "No Seam"
        elif seam == "Side Seam":
            seam = "Side Seam"

        sam_data.append({
            'gender': str(a_val).strip(),
            'product': str(b_val).strip(),
            'seam': seam,
            'size': str(d_val).strip(),
            'sam_minutes': float(e_val)
        })

print(f"Found {len(sam_data)} rows")

# Write to CSV
csv_path = OUT_DIR / "sam_minutes_lookup.csv"
with csv_path.open("w", newline="", encoding="utf-8") as f:
    w = csv.DictWriter(f, fieldnames=['gender', 'product', 'seam', 'size', 'sam_minutes'])
    w.writeheader()
    for row in sam_data:
        w.writerow(row)

print(f"Wrote {len(sam_data)} rows to {csv_path}")

# Verify some key combinations
print("\n" + "=" * 80)
print("VERIFICATION - Key Combinations")
print("=" * 80)

test_cases = [
    ("Men", "Tank top/A Shirt", "Side Seam", "S-XL"),
    ("Men", "T-Shirt (Crew Neck)", "No Seam", "S-XL"),
    ("Men", "Long Sleeve Shirt (Crew Neck)", "Side Seam", "S-XL"),
]

for gender, product, seam, size in test_cases:
    for row in sam_data:
        if (row['gender'] == gender and
            row['product'] == product and
            row['seam'] == seam and
            row['size'] == size):
            print(f"{gender} + {product} + {seam} + {size} = {row['sam_minutes']} minutes")
            break
