"""
Extract ALL Total Cost Summary calculation rules from Excel
"""
import openpyxl

SRC = r"C:\Users\dploy\OneDrive\Documents\Design to Basic Shirt Tool First Trail Version 1 - 20250801-update latest.xlsm"

wb = openpyxl.load_workbook(SRC, data_only=False, keep_vba=True)
ws_dl = wb['Data link']

print("=" * 80)
print("EXTRACTING ALL COST CALCULATION RULES")
print("=" * 80)

# 1. Sewing Thread Cost (Row 26)
print("\n" + "=" * 80)
print("1. SEWING THREAD COST (Data link!X26)")
print("=" * 80)

print("\nFormula:")
print(ws_dl.cell(26, 24).value)  # Column X

print("\nRule: Cost depends on Silhouette selected")
print("  - Tank Top/A Shirt: $0.035")
print("  - T-Shirt (Crew Neck): $0.042")
print("  - T-Shirt (V Neck): T7 * 0.042")
print("  - Long Sleeve Shirt (Crew Neck): $0.048")
print("  - Long Sleeve Shirt (V Neck): V7 * 0.048")
print("  - Sleeveless Shirt (Crew Neck): $0.038")
print("  - Sleeveless Shirt (V Neck): $0.038")

# 2. Product Testing Cost (Row 25)
print("\n" + "=" * 80)
print("2. PRODUCT TESTING COST (Data link!X25)")
print("=" * 80)

print("\nFormula:")
print(ws_dl.cell(25, 24).value)  # Column X

# 3. Check what T7 and V7 are
print("\n" + "=" * 80)
print("CHECKING REFERENCES T7, V7")
print("=" * 80)

print("\nT7 (Data link):")
print(f"  Value: {ws_dl.cell(7, 20).value}")  # Column T

print("\nV7 (Data link):")
print(f"  Value: {ws_dl.cell(7, 22).value}")  # Column V

# 4. Check Basic Shirts for other references
print("\n" + "=" * 80)
print("BASIC SHIRTS - CHECKING ALL TOTAL COST REFERENCES")
print("=" * 80)

ws_bs = wb['Basic Shirts Costing Tool']

for row in range(6, 23):
    label = ws_bs.cell(row, 27).value  # AA column
    if not label:
        continue

    ac_formula = ws_bs.cell(row, 29).value  # AC column formula

    print(f"\n{label}")
    print(f"  AC Formula: {ac_formula}")

# 5. Check for overhead costs
print("\n" + "=" * 80)
print("CHECKING FOR OVERHEAD COSTS")
print("=" * 80)

print("\nData link Row 25 (Product Testing):")
for col in range(1, 30):
    val = ws_dl.cell(25, col).value
    if val is not None:
        col_letter = openpyxl.utils.get_column_letter(col)
        print(f"  {col_letter}25: {val}")
