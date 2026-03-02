"""
Check what the screenshot is actually showing
"""
import openpyxl

SRC = r"C:\Users\dploy\Downloads\Design to Basic Shirt Tool First Trail Version 1 - 20250801-update latest.xlsm"

wb = openpyxl.load_workbook(SRC, data_only=True, keep_vba=True)
ws_bs = wb['Basic Shirts Costing Tool']

print("=" * 80)
print("MANUFACTURING COST SECTION - What Excel Actually Shows")
print("=" * 80)

print("\n=== Rows 18-21 (Manufacturing section) ===")
for row in range(18, 22):
    label = ws_bs.cell(row, 19).value  # S column
    val_w = ws_bs.cell(row, 23).value  # W column
    val_y = ws_bs.cell(row, 25).value  # Y column
    if label:
        print(f"Row {row}: {label}")
        print(f"  W{row} = {val_w}")
        print(f"  Y{row} = {val_y}")

print("\n=== AC17 (Total Cost Summary) ===")
ac17_val = ws_bs.cell(17, 29).value
ac17_label = ws_bs.cell(17, 27).value
print(f"AC17 ({ac17_label}): {ac17_val}")

print("\n" + "=" * 80)
print("QUESTION: Which cell shows 2.3 in your screenshot?")
print("  - W21 (Total making cost per garment)?")
print("  - AC17 (SUPPLIER FOB COST PER PIECE)?")
print("=" * 80)
