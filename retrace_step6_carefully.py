"""
Carefully re-trace Step 6 Manufacturing Cost from screenshots
"""
import openpyxl

SRC = r"C:\Users\dploy\Downloads\Design to Basic Shirt Tool First Trail Version 1 - 20250801-update latest.xlsm"

wb_f = openpyxl.load_workbook(SRC, data_only=False, keep_vba=True)
wb_v = openpyxl.load_workbook(SRC, data_only=True, keep_vba=True)

ws_bs_f = wb_f['Basic Shirts Costing Tool']
ws_bs_v = wb_v['Basic Shirts Costing Tool']
ws_dl_f = wb_f['Data link']
ws_dl_v = wb_v['Data link']

print("=" * 80)
print("STEP 6: MANUFACTURING COST - CORRECT TRACE")
print("=" * 80)

print("\n=== BASIC SHIRTS - Manufacturing Section (W18-W21) ===")
for row in range(18, 22):
    label = ws_bs_v.cell(row, 19).value  # S column
    w_val = ws_bs_v.cell(row, 23).value  # W column
    w_formula = ws_bs_f.cell(row, 23).value
    print(f"\nRow {row}: {label}")
    print(f"  W{row} Value: {w_val}")
    print(f"  W{row} Formula: {w_formula}")

print("\n" + "=" * 80)
print("DATA LINK - Row 30 (Source for W18-W20)")
print("=" * 80)

# Get all relevant columns
cols = {
    'J': 'Base minutes before pocket bag',
    'K': 'Minutes after pocket bag check',
    'L': 'Cost Rate',
    'N': 'Efficiency multiplier',
    'Q': 'Efficiency',
    'R': 'Total (basic)',
}

for col, label in cols.items():
    val = ws_dl_v[f'{col}30'].value
    formula = ws_dl_f[f'{col}30'].value
    print(f"\n{col}30 ({label}):")
    print(f"  Value: {val}")
    print(f"  Formula: {formula}")

print("\n" + "=" * 80)
print("WHAT ARE THE CORRECT VALUES FROM YOUR EXCEL SCREENSHOT?")
print("=" * 80)
print("You said results show 2.3, but I'm getting different values.")
print("Please tell me:")
print("1. Which cell in Excel shows 2.3?")
print("2. What is the exact label/row number?")
print("3. Is it in Basic Shirts or Data link sheet?")
print("=" * 80)
