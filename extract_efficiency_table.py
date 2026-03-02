"""
Extract efficiency lookup table
"""
import csv
from pathlib import Path
import openpyxl

SRC = r"C:\Users\dploy\OneDrive\Documents\Design to Basic Shirt Tool First Trail Version 1 - 20250801-update latest.xlsm"
OUT_DIR = Path(r"C:\Users\dploy\.openclaw\workspace\basicshirts_web\master_clean")

wb = openpyxl.load_workbook(SRC, data_only=True, keep_vba=True)
ws_sam = wb['SAM&Product EFF%']

print("=" * 80)
print("EFFICIENCY LOOKUP TABLE")
print("=" * 80)

efficiency_data = []

for row in range(3, 10):  # Start from row 3 (skip headers)
    quantity = ws_sam.cell(row, 27).value  # Column AA
    efficiency = ws_sam.cell(row, 28).value  # Column AB

    if quantity and efficiency is not None:
        efficiency_data.append({
            'quantity_range': str(quantity).strip(),
            'efficiency': float(efficiency)
        })
        print(f"{quantity} -> {efficiency}")

# Write to CSV
csv_path = OUT_DIR / "efficiency_by_quantity.csv"
with csv_path.open("w", newline="", encoding="utf-8") as f:
    w = csv.DictWriter(f, fieldnames=['quantity_range', 'efficiency'])
    w.writeheader()
    for row in efficiency_data:
        w.writerow(row)

print(f"\nWrote {len(efficiency_data)} rows to {csv_path}")
