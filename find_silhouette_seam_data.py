"""
Check what silhouette+seam combination gives 2.3 minutes
"""
import openpyxl

SRC = r"C:\Users\dploy\Downloads\Design to Basic Shirt Tool First Trail Version 1 - 20250801-update latest.xlsm"

wb = openpyxl.load_workbook(SRC, data_only=True, keep_vba=True)

ws_sam = wb['SAM&Product EFF%']

print("=" * 80)
print("SAM&Product EFF% SHEET - ALL DATA")
print("=" * 80)

# Check all columns to find minutes data
for row in range(1, 50):
    row_data = []
    for col in range(1, 30):
        val = ws_sam.cell(row, col).value
        if val is not None:
            col_letter = openpyxl.utils.get_column_letter(col)
            row_data.append(f"{col_letter}={val}")

    if row_data:
        print(f"Row {row}: {', '.join(row_data[:5])}")  # Show first 5 values

print("\n" + "=" * 80)
print("LOOKING FOR SILHOUETTE+SEAM COMBINATIONS")
print("=" * 80)

# Look for silhouette names in the sheet
for row in range(1, 50):
    for col in range(1, 30):
        val = ws_sam.cell(row, col).value
        if val and isinstance(val, str) and ('sleeve' in val.lower() or 'shirt' in val.lower() or 'seam' in val.lower()):
            col_letter = openpyxl.utils.get_column_letter(col)
            print(f"Row {row}, {col_letter}{row}: {val}")
