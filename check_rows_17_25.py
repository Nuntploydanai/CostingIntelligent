"""
Check rows 17-25 for additional cost items
"""
import openpyxl

SRC = r"C:\Users\dploy\OneDrive\Documents\Design to Basic Shirt Tool First Trail Version 1 - 20250801-update latest.xlsm"

wb = openpyxl.load_workbook(SRC, data_only=True, keep_vba=True)
ws_bs = wb['Basic Shirts Costing Tool']

print("=" * 80)
print("BASIC SHIRTS COSTING TOOL - ROWS 17-25")
print("=" * 80)

print(f"\n{'Row':<5} {'AA (Label)':<50} {'AC (Value)':<20}")
print("-" * 80)

for row in range(17, 30):
    label = ws_bs.cell(row, 27).value  # Column AA
    value = ws_bs.cell(row, 29).value  # Column AC

    if label or value is not None:
        label_str = str(label)[:48] if label else ''
        value_str = str(value)[:18] if value is not None else ''
        print(f"{row:<5} {label_str:<50} {value_str:<20}")

# Check formulas
print("\n" + "=" * 80)
print("FORMULAS")
print("=" * 80)

wb2 = openpyxl.load_workbook(SRC, data_only=False, keep_vba=True)
ws_bs2 = wb2['Basic Shirts Costing Tool']

for row in range(17, 30):
    label = ws_bs2.cell(row, 27).value  # Column AA
    formula = ws_bs2.cell(row, 29).value  # Column AC

    if label:
        print(f"\nRow {row}: {label}")
        print(f"  Formula: {formula}")
