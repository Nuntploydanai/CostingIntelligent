"""
Check Data link inputs in user's Excel
"""
import openpyxl

SRC = r"C:\Users\dploy\OneDrive\Documents\Design to Basic Shirt Tool First Trail Version 1 - 20250801-update latest.xlsm"

wb = openpyxl.load_workbook(SRC, data_only=True, keep_vba=True)

ws_dl = wb['Data link']

print("=" * 80)
print("DATA LINK INPUTS (E25, E26, E27, E29)")
print("=" * 80)

print(f"E25 (Gender): {ws_dl['E25'].value}")
print(f"E26 (Silhouette): {ws_dl['E26'].value}")
print(f"E27 (Seam): {ws_dl['E27'].value}")
print(f"E29 (Size): {ws_dl['E29'].value}")

print(f"\nJ30 (Minutes from lookup): {ws_dl['J30'].value}")
print(f"K30 (Minutes with pocket adjustment): {ws_dl['K30'].value}")
print(f"W18 (Final minutes): {ws_dl['W18'].value}")

print("\n" + "=" * 80)
print("Looking for 2.3 in SAM sheet")
print("=" * 80)

ws_sam = wb['SAM&Product EFF%']

for row_num in range(2, 100):
    e_val = ws_sam.cell(row_num, 5).value
    if e_val is not None:
        try:
            sam = float(e_val)
            if abs(sam - 2.3) < 0.01:
                a_val = ws_sam.cell(row_num, 1).value
                b_val = ws_sam.cell(row_num, 2).value
                c_val = ws_sam.cell(row_num, 3).value
                d_val = ws_sam.cell(row_num, 4).value
                print(f"Row {row_num}: {a_val} + {b_val} + {c_val} + {d_val} = {sam}")
        except:
            pass
