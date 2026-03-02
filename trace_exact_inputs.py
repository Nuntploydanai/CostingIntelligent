"""
Trace with EXACT inputs: T-Shirt (Crew Neck) + No Seam
"""
import openpyxl

SRC = r"C:\Users\dploy\Downloads\Design to Basic Shirt Tool First Trail Version 1 - 20250801-update latest.xlsm"

wb_f = openpyxl.load_workbook(SRC, data_only=False, keep_vba=True)
wb_v = openpyxl.load_workbook(SRC, data_only=True, keep_vba=True)

ws_dl_f = wb_f['Data link']
ws_dl_v = wb_v['Data link']

print("=" * 80)
print("EXACT INPUTS FROM SCREENSHOT")
print("=" * 80)
print("Silhouette: T-Shirt (Crew Neck)")
print("Seam: No Seam")
print("Ideal Quantity: More than 100,000")

print("\n" + "=" * 80)
print("DATA LINK - Find row for T-Shirt (Crew Neck) + No Seam")
print("=" * 80)

# Find the row in Data link that has this combination
for row in range(1, 100):
    b_val = ws_dl_v.cell(row, 2).value  # B column - silhouette
    c_val = ws_dl_v.cell(row, 3).value  # C column - seam
    d_val = ws_dl_v.cell(row, 4).value  # D column - key

    if b_val and 'T-Shirt' in str(b_val) and 'Crew' in str(b_val):
        if c_val and 'No Seam' in str(c_val):
            print(f"\nFound at Row {row}:")
            print(f"  B{row} (Silhouette): {b_val}")
            print(f"  C{row} (Seam): {c_val}")
            print(f"  D{row} (Key): {d_val}")

            # Get minutes and other values
            j_val = ws_dl_v.cell(row, 10).value  # J column
            k_val = ws_dl_v.cell(row, 11).value  # K column

            print(f"  J{row} (Base Minutes): {j_val}")
            print(f"  K{row} (Minutes): {k_val}")

            # Check formulas
            j_formula = ws_dl_f.cell(row, 10).value
            k_formula = ws_dl_f.cell(row, 11).value

            print(f"  J{row} Formula: {j_formula}")
            print(f"  K{row} Formula: {k_formula}")

print("\n" + "=" * 80)
print("EXPECTED: W18 = 2.3")
print("Need to find which row gives 2.3 minutes")
print("=" * 80)
