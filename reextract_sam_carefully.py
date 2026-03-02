"""
Re-extract SAM data - more carefully this time
"""
import openpyxl
import csv
from pathlib import Path

SRC = r"C:\Users\dploy\Downloads\Design to Basic Shirt Tool First Trail Version 1 - 20250801-update latest.xlsm"
OUT_DIR = Path(r"C:\Users\dploy\.openclaw\workspace\basicshirts_web\master_clean")

wb = openpyxl.load_workbook(SRC, data_only=True, keep_vba=True)
ws_sam = wb['SAM&Product EFF%']

print("=" * 80)
print("RE-EXTRACTING SAM DATA - VERIFYING EACH ROW")
print("=" * 80)

# First, check headers
print("\n=== Headers (Row 1) ===")
for col in range(1, 10):
    val = ws_sam.cell(1, col).value
    if val:
        col_letter = openpyxl.utils.get_column_letter(col)
        print(f"{col_letter}1: {val}")

print("\n=== Row 20 (should be Men + T-Shirt (Crew Neck) + No seam) ===")
for col in range(1, 10):
    val = ws_sam.cell(20, col).value
    col_letter = openpyxl.utils.get_column_letter(col)
    print(f"{col_letter}20: {val}")

print("\n=== Extracting all data ===")
sam_data = []

for row in range(2, 200):
    a_val = ws_sam.cell(row, 1).value  # Gender
    b_val = ws_sam.cell(row, 2).value  # Product shape
    c_val = ws_sam.cell(row, 3).value  # Seam
    d_val = ws_sam.cell(row, 4).value  # Size
    e_val = ws_sam.cell(row, 5).value  # SAM (minutes)

    if a_val and b_val and c_val and d_val and e_val is not None:
        sam_data.append({
            'gender': str(a_val).strip(),
            'product': str(b_val).strip(),
            'seam': str(c_val).strip(),
            'size': str(d_val).strip(),
            'sam_minutes': float(e_val)
        })

print(f"Found {len(sam_data)} rows")

# Check T-Shirt (Crew Neck) + No seam
print("\n=== T-Shirt (Crew Neck) + No seam combinations ===")
for row in sam_data:
    if 'T-Shirt (Crew Neck)' in row['product'] and 'No seam' in row['seam']:
        print(f"{row['gender']} + {row['product']} + {row['seam']} + {row['size']} = {row['sam_minutes']}")

# Write to CSV
OUT_DIR.mkdir(parents=True, exist_ok=True)
with (OUT_DIR / "sam_minutes_lookup.csv").open("w", newline="", encoding="utf-8") as f:
    w = csv.DictWriter(f, fieldnames=['gender', 'product', 'seam', 'size', 'sam_minutes'])
    w.writeheader()
    for row in sam_data:
        w.writerow(row)

print(f"\nWrote {len(sam_data)} rows to sam_minutes_lookup.csv")
