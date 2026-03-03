"""
CAREFULLY trace ALL Total Cost Summary values from Excel
"""
import openpyxl

SRC = r"C:\Users\dploy\OneDrive\Documents\Design to Basic Shirt Tool First Trail Version 1 - 20250801-update latest.xlsm"

wb = openpyxl.load_workbook(SRC, data_only=False, keep_vba=True)
ws_dl = wb['Data link']

print("=" * 80)
print("TRACING TOTAL COST SUMMARY - CELL BY CELL")
print("=" * 80)

# Check row 26 (sewing thread row mentioned by user)
print("\n" + "=" * 80)
print("ROW 26 - SEWING THREAD")
print("=" * 80)

for col in range(1, 30):
    val = ws_dl.cell(26, col).value
    if val is not None:
        col_letter = openpyxl.utils.get_column_letter(col)
        print(f"{col_letter}26: {val}")

# Check the actual formulas for AC column (rows 2-15)
print("\n" + "=" * 80)
print("ALL TOTAL COST FORMULAS (Column AC)")
print("=" * 80)

wb_data = openpyxl.load_workbook(SRC, data_only=True, keep_vba=True)
ws_dl_data = wb_data['Data link']

for row in range(2, 16):
    label = ws_dl.cell(row, 27).value  # AA
    if not label:
        continue

    # Get formula from AC
    ac_formula = ws_dl.cell(row, 29).value  # Column AC formula
    ac_value = ws_dl_data.cell(row, 29).value  # Column AC value

    print(f"\n{label}")
    print(f"  Formula: {ac_formula}")
    print(f"  Value: {ac_value}")

# Now check Basic Shirts sheet for the source of these values
print("\n" + "=" * 80)
print("BASIC SHIRTS - SOURCE OF VALUES")
print("=" * 80)

ws_bs = wb['Basic Shirts Costing Tool']
ws_bs_data = wb_data['Basic Shirts Costing Tool']

for row in range(6, 20):
    label = ws_bs.cell(row, 27).value  # AA
    if not label:
        continue

    # Get formula from AC
    ac_formula = ws_bs.cell(row, 29).value
    ac_value = ws_bs_data.cell(row, 29).value

    print(f"\n{label}")
    print(f"  Formula: {ac_formula}")
    print(f"  Value: {ac_value}")

# Check for sewing thread specifically
print("\n" + "=" * 80)
print("SEWING THREAD DETAILED TRACE")
print("=" * 80)

# Check if there's a specific cell for sewing thread
print("\nLooking for sewing thread references...")
for sheet_name in ['Data link', 'Basic Shirts Costing Tool']:
    ws = wb[sheet_name]
    print(f"\n{sheet_name}:")

    for row in range(1, 50):
        for col in range(1, 30):
            val = ws.cell(row, col).value
            if val and isinstance(val, str) and 'thread' in val.lower():
                col_letter = openpyxl.utils.get_column_letter(col)
                print(f"  {col_letter}{row}: {val}")

                # Check next few columns for values
                for next_col in range(col + 1, col + 5):
                    next_val = ws.cell(row, next_col).value
                    if next_val is not None:
                        next_col_letter = openpyxl.utils.get_column_letter(next_col)
                        print(f"    {next_col_letter}{row}: {next_val}")
