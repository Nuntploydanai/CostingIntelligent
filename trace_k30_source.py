"""
Check Data link K30 with different silhouette+seam
"""
import openpyxl

SRC = r"C:\Users\dploy\Downloads\Design to Basic Shirt Tool First Trail Version 1 - 20250801-update latest.xlsm"

wb_f = openpyxl.load_workbook(SRC, data_only=False, keep_vba=True)
wb_v = openpyxl.load_workbook(SRC, data_only=True, keep_vba=True)

ws_dl_f = wb_f['Data link']
ws_dl_v = wb_v['Data link']

print("=" * 80)
print("DATA LINK K30 - THE SOURCE OF W18")
print("=" * 80)

print("\nK30 Formula:")
print(ws_dl_f['K30'].value)

print("\nK30 Value:")
print(ws_dl_v['K30'].value)

print("\nJ30 Formula (base minutes):")
print(ws_dl_f['J30'].value)

print("\nJ30 Value:")
print(ws_dl_v['J30'].value)

print("\n" + "=" * 80)
print("J30 is an ARRAY FORMULA")
print("It likely looks up based on E7 (silhouette+seam)")
print("=" * 80)

print("\nE7 (silhouette+seam key):")
print(ws_dl_v['E7'].value)

# Check if J30 depends on silhouette
print("\n" + "=" * 80)
print("CHECK ALL COLUMNS IN ROW 7")
print("=" * 80)
for col in range(1, 30):
    val_v = ws_dl_v.cell(7, col).value
    val_f = ws_dl_f.cell(7, col).value
    if val_v is not None or val_f is not None:
        col_letter = openpyxl.utils.get_column_letter(col)
        print(f"{col_letter}7: {val_v} | Formula: {val_f}")
