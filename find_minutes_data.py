"""
Find where silhouette+seam minutes are stored
"""
import openpyxl

SRC = r"C:\Users\dploy\Downloads\Design to Basic Shirt Tool First Trail Version 1 - 20250801-update latest.xlsm"

wb = openpyxl.load_workbook(SRC, data_only=True, keep_vba=True)

ws_sam = wb['SAM&Product EFF%']

print("=" * 80)
print("SAM&Product EFF% SHEET - Looking for minutes data")
print("=" * 80)

# Check first 50 rows, all columns
print("\nFirst row with data:")
for row in range(1, 10):
    row_data = {}
    for col in range(1, 50):
        val = ws_sam.cell(row, col).value
        if val is not None:
            col_letter = openpyxl.utils.get_column_letter(col)
            row_data[col_letter] = val

    if row_data:
        print(f"\nRow {row}:")
        for col, val in row_data.items():
            print(f"  {col}: {val}")

print("\n" + "=" * 80)
print("Look for silhouette names and minutes")
print("=" * 80)

# Search for T_Shirt_No_Seam or similar keys
for row in range(1, 200):
    for col in range(1, 50):
        val = ws_sam.cell(row, col).value
        if val and isinstance(val, str):
            if 'T_Shirt' in val or 't-shirt' in val.lower():
                col_letter = openpyxl.utils.get_column_letter(col)
                print(f"Row {row}, {col_letter}{row}: {val}")

                # Print nearby values
                for c in range(max(1, col-2), min(50, col+3)):
                    v = ws_sam.cell(row, c).value
                    if v is not None:
                        cl = openpyxl.utils.get_column_letter(c)
                        print(f"  {cl}{row}: {v}")
